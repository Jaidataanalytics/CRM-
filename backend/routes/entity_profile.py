from fastapi import APIRouter, HTTPException, Request, Depends, Query
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import logging

from models.user import User
from routes.auth import get_current_user

router = APIRouter(prefix="/entity", tags=["Entity Profile"])
logger = logging.getLogger(__name__)


async def get_db(request: Request):
    return request.app.state.db


def get_indian_fy_dates():
    """Get current Indian Financial Year dates (April 1 - March 31)"""
    today = datetime.now()
    if today.month >= 4:
        start_year = today.year
    else:
        start_year = today.year - 1
    
    start_date = f"{start_year}-04-01"
    end_date = f"{start_year + 1}-03-31"
    return start_date, end_date


async def get_metric_config(db, metric_id: str) -> dict:
    """Get metric configuration from database"""
    metric = await db.metric_settings.find_one({"metric_id": metric_id}, {"_id": 0})
    return metric


async def count_by_metric(db, base_query: dict, metric_config: dict) -> int:
    """Count leads matching a metric configuration"""
    if not metric_config or not metric_config.get("is_active", True):
        return 0
    
    field_name = metric_config.get("field_name")
    field_values = metric_config.get("field_values", [])
    
    if not field_name or not field_values:
        return 0
    
    # Handle boolean fields
    if len(field_values) == 1 and isinstance(field_values[0], bool):
        query = {**base_query, field_name: field_values[0]}
    else:
        query = {**base_query, field_name: {"$in": field_values}}
    
    # Exclude soft-deleted leads
    query["is_deleted"] = {"$ne": True}
    
    return await db.leads.count_documents(query)


@router.get("/search")
async def search_entities(
    request: Request,
    q: str = Query(..., min_length=2),
    current_user: User = Depends(get_current_user)
):
    """Search for states, dealers, cities, and employees"""
    db = await get_db(request)
    results = []
    query_lower = q.lower()
    
    # Search states
    states = await db.leads.distinct("state")
    for state in states:
        if state and query_lower in state.lower():
            count = await db.leads.count_documents({"state": state})
            results.append({
                "type": "state",
                "name": state,
                "id": state,
                "lead_count": count
            })
    
    # Search dealers
    dealers = await db.leads.distinct("dealer")
    for dealer in dealers:
        if dealer and query_lower in dealer.lower():
            # Get dealer's state
            sample = await db.leads.find_one({"dealer": dealer})
            count = await db.leads.count_documents({"dealer": dealer})
            results.append({
                "type": "dealer",
                "name": dealer,
                "id": dealer,
                "state": sample.get("state") if sample else None,
                "lead_count": count
            })
    
    # Search cities/areas
    cities = await db.leads.distinct("area")
    for city in cities:
        if city and query_lower in city.lower():
            sample = await db.leads.find_one({"area": city})
            count = await db.leads.count_documents({"area": city})
            results.append({
                "type": "city",
                "name": city,
                "id": city,
                "state": sample.get("state") if sample else None,
                "lead_count": count
            })
    
    # Search employees
    employees = await db.leads.distinct("employee_name")
    for emp in employees:
        if emp and query_lower in emp.lower():
            sample = await db.leads.find_one({"employee_name": emp})
            count = await db.leads.count_documents({"employee_name": emp})
            results.append({
                "type": "employee",
                "name": emp,
                "id": emp,
                "dealer": sample.get("dealer") if sample else None,
                "state": sample.get("state") if sample else None,
                "lead_count": count
            })
    
    # Sort by lead count descending
    results.sort(key=lambda x: x.get("lead_count", 0), reverse=True)
    
    return {"results": results[:20]}


@router.get("/profile/{entity_type}/{entity_id}")
async def get_entity_profile(
    request: Request,
    entity_type: str,
    entity_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get comprehensive profile for an entity (state/dealer/city/employee)"""
    db = await get_db(request)
    
    if entity_type not in ["state", "dealer", "city", "employee"]:
        raise HTTPException(status_code=400, detail="Invalid entity type")
    
    # Build filter based on entity type
    field_map = {
        "state": "state",
        "dealer": "dealer",
        "city": "area",
        "employee": "employee_name"
    }
    
    filter_field = field_map[entity_type]
    
    # Use Indian FY dates as default if no dates provided
    if not start_date or not end_date:
        start_date, end_date = get_indian_fy_dates()
    
    # Base filter with entity + date range (same as dashboard)
    base_filter = {
        filter_field: entity_id,
        "is_deleted": {"$ne": True},
        "enquiry_date": {
            "$gte": start_date,
            "$lte": end_date
        }
    }
    
    # Check if entity exists (without date filter)
    exists = await db.leads.find_one({filter_field: entity_id, "is_deleted": {"$ne": True}})
    if not exists:
        raise HTTPException(status_code=404, detail=f"{entity_type.capitalize()} not found")
    
    # Get basic info from sample lead
    sample_lead = await db.leads.find_one({filter_field: entity_id, "is_deleted": {"$ne": True}})
    
    # Build profile data
    profile = {
        "entity_type": entity_type,
        "entity_id": entity_id,
        "entity_name": entity_id,
        "date_range": {
            "start_date": start_date,
            "end_date": end_date
        }
    }
    
    # Add parent info based on entity type
    if entity_type == "dealer":
        profile["state"] = sample_lead.get("state")
    elif entity_type == "city":
        profile["state"] = sample_lead.get("state")
    elif entity_type == "employee":
        profile["dealer"] = sample_lead.get("dealer")
        profile["state"] = sample_lead.get("state")
    
    # Get KPIs using admin-configurable metrics (same as dashboard)
    won_config = await get_metric_config(db, "won_leads")
    lost_config = await get_metric_config(db, "lost_leads")
    open_config = await get_metric_config(db, "open_leads")
    closed_config = await get_metric_config(db, "closed_leads")
    hot_config = await get_metric_config(db, "hot_leads")
    warm_config = await get_metric_config(db, "warm_leads")
    cold_config = await get_metric_config(db, "cold_leads")
    
    total_leads = await db.leads.count_documents(base_filter)
    won_leads = await count_by_metric(db, base_filter, won_config)
    lost_leads = await count_by_metric(db, base_filter, lost_config)
    open_leads = await count_by_metric(db, base_filter, open_config)
    closed_leads = await count_by_metric(db, base_filter, closed_config)
    hot_leads = await count_by_metric(db, base_filter, hot_config)
    warm_leads = await count_by_metric(db, base_filter, warm_config)
    cold_leads = await count_by_metric(db, base_filter, cold_config)
    
    # Call & Quotation metrics
    calls_placed_statuses = [
        'Called - No Response', 'Called - Interested', 
        'Called - Not Interested', 'Called - Follow Up Required', 'Called - Converted'
    ]
    calls_placed = await db.leads.count_documents({**base_filter, "call_status": {"$in": calls_placed_statuses}})
    not_called = await db.leads.count_documents({
        **base_filter, 
        "$or": [{"call_status": {"$exists": False}}, {"call_status": None}, {"call_status": "Not Called"}]
    })
    quotations_sent = await db.leads.count_documents({**base_filter, "quotation_sent": True})
    
    # Conversion Rate = Won Leads / Total Leads (not just closed leads)
    conversion_rate = (won_leads / total_leads * 100) if total_leads > 0 else 0
    call_to_quotation_rate = (quotations_sent / calls_placed * 100) if calls_placed > 0 else 0
    
    profile["kpis"] = {
        "total_leads": total_leads,
        "open_leads": open_leads,
        "won_leads": won_leads,
        "lost_leads": lost_leads,
        "closed_leads": closed_leads,
        "hot_leads": hot_leads,
        "warm_leads": warm_leads,
        "cold_leads": cold_leads,
        "conversion_rate": round(conversion_rate, 2),
        "calls_placed": calls_placed,
        "not_called": not_called,
        "quotations_sent": quotations_sent,
        "call_to_quotation_rate": round(call_to_quotation_rate, 2)
    }
    
    # Avg Lead Age (for open leads)
    pipeline_age = [
        {"$match": {**base_filter, "enquiry_stage": {"$in": ["Prospecting", "Qualified", "Proposal", "Negotiation"]}}},
        {"$addFields": {
            "enquiry_date_parsed": {"$dateFromString": {"dateString": "$enquiry_date", "onError": None}}
        }},
        {"$match": {"enquiry_date_parsed": {"$ne": None}}},
        {"$project": {
            "age_days": {
                "$divide": [
                    {"$subtract": [datetime.now(timezone.utc), "$enquiry_date_parsed"]},
                    1000 * 60 * 60 * 24
                ]
            }
        }},
        {"$group": {"_id": None, "avg_age": {"$avg": "$age_days"}}}
    ]
    
    age_result = await db.leads.aggregate(pipeline_age).to_list(1)
    avg_lead_age = round(age_result[0]["avg_age"], 1) if age_result and age_result[0].get("avg_age") else 0
    
    # Avg Closure Time (for closed leads)
    pipeline_closure = [
        {"$match": {**base_filter, "enquiry_stage": {"$in": ["Closed-Won", "Order Booked", "Closed-Lost", "Closed-Dropped"]}}},
        {"$addFields": {
            "enquiry_date_parsed": {"$dateFromString": {"dateString": "$enquiry_date", "onError": None}},
            "followup_date_parsed": {"$dateFromString": {"dateString": "$last_followup_date", "onError": None}}
        }},
        {"$match": {"enquiry_date_parsed": {"$ne": None}, "followup_date_parsed": {"$ne": None}}},
        {"$project": {
            "closure_days": {
                "$divide": [
                    {"$subtract": ["$followup_date_parsed", "$enquiry_date_parsed"]},
                    1000 * 60 * 60 * 24
                ]
            }
        }},
        {"$match": {"closure_days": {"$gte": 0}}},
        {"$group": {"_id": None, "avg_closure": {"$avg": "$closure_days"}}}
    ]
    
    closure_result = await db.leads.aggregate(pipeline_closure).to_list(1)
    avg_closure_time = round(closure_result[0]["avg_closure"], 1) if closure_result and closure_result[0].get("avg_closure") else 0
    
    profile["kpis"]["avg_lead_age"] = avg_lead_age
    profile["kpis"]["avg_closure_time"] = avg_closure_time
    
    # Lead Source Breakdown
    source_pipeline = [
        {"$match": base_filter},
        {"$group": {"_id": "$source", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    source_breakdown = await db.leads.aggregate(source_pipeline).to_list(20)
    profile["source_breakdown"] = [{"source": s["_id"] or "Unknown", "count": s["count"]} for s in source_breakdown]
    
    # Duplicate Leads Count
    duplicate_count = await db.leads.count_documents({**base_filter, "is_duplicate": True})
    profile["duplicate_leads_count"] = duplicate_count
    
    # Order Time Punch - leads with sales order data
    order_punch_count = await db.leads.count_documents({
        **base_filter,
        "sales_order_no": {"$exists": True, "$ne": None, "$ne": ""},
        "sales_order_date": {"$exists": True, "$ne": None, "$ne": ""}
    })
    profile["order_time_punch_count"] = order_punch_count
    
    # Segment Performance
    segment_pipeline = [
        {"$match": base_filter},
        {"$group": {
            "_id": "$segment",
            "total": {"$sum": 1},
            "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", ["Closed-Won", "Order Booked"]]}, 1, 0]}},
            "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", ["Closed-Lost", "Closed-Dropped"]]}, 1, 0]}}
        }},
        {"$sort": {"total": -1}}
    ]
    segment_data = await db.leads.aggregate(segment_pipeline).to_list(20)
    profile["segment_performance"] = [{
        "segment": s["_id"] or "Unknown",
        "total": s["total"],
        "won": s["won"],
        "lost": s["lost"],
        "conversion_rate": round((s["won"] / s["total"] * 100), 1) if s["total"] > 0 else 0
    } for s in segment_data]
    
    # Follow-up Status
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    overdue_filter = {
        **base_filter,
        "planned_followup_date": {"$lt": today},
        "enquiry_stage": {"$nin": ["Closed-Won", "Order Booked", "Closed-Lost", "Closed-Dropped"]}
    }
    upcoming_filter = {
        **base_filter,
        "planned_followup_date": {"$gte": today},
        "enquiry_stage": {"$nin": ["Closed-Won", "Order Booked", "Closed-Lost", "Closed-Dropped"]}
    }
    
    overdue_count = await db.leads.count_documents(overdue_filter)
    upcoming_count = await db.leads.count_documents(upcoming_filter)
    
    profile["followup_status"] = {
        "overdue": overdue_count,
        "on_track": upcoming_count
    }
    
    # Trend Over Time (last 6 months)
    six_months_ago = (datetime.now(timezone.utc) - timedelta(days=180)).strftime("%Y-%m-%d")
    trend_pipeline = [
        {"$match": {filter_field: entity_id, "enquiry_date": {"$gte": six_months_ago}}},
        {"$addFields": {
            "month": {"$substr": ["$enquiry_date", 0, 7]}
        }},
        {"$group": {
            "_id": "$month",
            "total": {"$sum": 1},
            "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", ["Closed-Won", "Order Booked"]]}, 1, 0]}},
            "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", ["Closed-Lost", "Closed-Dropped"]]}, 1, 0]}}
        }},
        {"$sort": {"_id": 1}}
    ]
    trend_data = await db.leads.aggregate(trend_pipeline).to_list(12)
    profile["trend"] = [{
        "month": t["_id"],
        "total": t["total"],
        "won": t["won"],
        "lost": t["lost"]
    } for t in trend_data]
    
    # Month-over-Month Comparison
    current_month = datetime.now(timezone.utc).strftime("%Y-%m")
    prev_month = (datetime.now(timezone.utc).replace(day=1) - timedelta(days=1)).strftime("%Y-%m")
    
    current_month_filter = {filter_field: entity_id, "is_deleted": {"$ne": True}, "enquiry_date": {"$regex": f"^{current_month}"}}
    prev_month_filter = {filter_field: entity_id, "is_deleted": {"$ne": True}, "enquiry_date": {"$regex": f"^{prev_month}"}}
    
    current_count = await db.leads.count_documents(current_month_filter)
    prev_count = await db.leads.count_documents(prev_month_filter)
    
    mom_change = ((current_count - prev_count) / prev_count * 100) if prev_count > 0 else 0
    
    profile["mom_comparison"] = {
        "current_month": current_month,
        "current_count": current_count,
        "prev_month": prev_month,
        "prev_count": prev_count,
        "change_percent": round(mom_change, 1)
    }
    
    # Get won/lost field values from metric configs for sub-entity calculations
    won_stages = won_config.get("field_values", ["Closed-Won", "Order Booked"]) if won_config else ["Closed-Won", "Order Booked"]
    lost_stages = lost_config.get("field_values", ["Closed-Lost", "Closed-Dropped"]) if lost_config else ["Closed-Lost", "Closed-Dropped"]
    
    # Sub-entities based on entity type (WITH date filter)
    if entity_type == "state":
        # Dealers in this state
        dealer_pipeline = [
            {"$match": base_filter},
            {"$group": {
                "_id": "$dealer",
                "total": {"$sum": 1},
                "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", won_stages]}, 1, 0]}},
                "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", lost_stages]}, 1, 0]}}
            }},
            {"$sort": {"total": -1}},
            {"$limit": 20}
        ]
        dealers = await db.leads.aggregate(dealer_pipeline).to_list(20)
        profile["sub_entities"] = {
            "dealers": [{
                "name": d["_id"],
                "total": d["total"],
                "won": d["won"],
                "lost": d["lost"],
                "conversion_rate": round((d["won"] / (d["won"] + d["lost"]) * 100), 1) if (d["won"] + d["lost"]) > 0 else 0
            } for d in dealers if d["_id"]]
        }
        
        # Districts in this state
        district_pipeline = [
            {"$match": base_filter},
            {"$group": {
                "_id": "$district",
                "total": {"$sum": 1},
                "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", won_stages]}, 1, 0]}},
                "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", lost_stages]}, 1, 0]}}
            }},
            {"$sort": {"won": -1}},
            {"$limit": 20}
        ]
        districts = await db.leads.aggregate(district_pipeline).to_list(20)
        profile["sub_entities"]["districts"] = [{
            "name": d["_id"],
            "total": d["total"],
            "won": d["won"],
            "lost": d["lost"],
            "conversion_rate": round((d["won"] / (d["won"] + d["lost"]) * 100), 1) if (d["won"] + d["lost"]) > 0 else 0
        } for d in districts if d["_id"]]
        
    elif entity_type == "dealer":
        # Employees under this dealer
        emp_pipeline = [
            {"$match": base_filter},
            {"$group": {
                "_id": "$employee_name",
                "total": {"$sum": 1},
                "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", won_stages]}, 1, 0]}},
                "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", lost_stages]}, 1, 0]}}
            }},
            {"$sort": {"total": -1}},
            {"$limit": 20}
        ]
        employees = await db.leads.aggregate(emp_pipeline).to_list(20)
        profile["sub_entities"] = {
            "employees": [{
                "name": e["_id"],
                "total": e["total"],
                "won": e["won"],
                "lost": e["lost"],
                "conversion_rate": round((e["won"] / (e["won"] + e["lost"]) * 100), 1) if (e["won"] + e["lost"]) > 0 else 0
            } for e in employees if e["_id"]]
        }
        
        # Districts covered by this dealer
        district_pipeline = [
            {"$match": base_filter},
            {"$group": {
                "_id": "$district",
                "total": {"$sum": 1},
                "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", won_stages]}, 1, 0]}},
                "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", lost_stages]}, 1, 0]}}
            }},
            {"$sort": {"won": -1}},
            {"$limit": 20}
        ]
        districts = await db.leads.aggregate(district_pipeline).to_list(20)
        profile["sub_entities"]["districts"] = [{
            "name": d["_id"],
            "total": d["total"],
            "won": d["won"],
            "lost": d["lost"],
            "conversion_rate": round((d["won"] / (d["won"] + d["lost"]) * 100), 1) if (d["won"] + d["lost"]) > 0 else 0
        } for d in districts if d["_id"]]
        
    elif entity_type == "city":
        # Dealers in this city
        dealer_pipeline = [
            {"$match": base_filter},
            {"$group": {
                "_id": "$dealer",
                "total": {"$sum": 1},
                "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", won_stages]}, 1, 0]}},
                "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", lost_stages]}, 1, 0]}}
            }},
            {"$sort": {"total": -1}},
            {"$limit": 20}
        ]
        dealers = await db.leads.aggregate(dealer_pipeline).to_list(20)
        profile["sub_entities"] = {
            "dealers": [{
                "name": d["_id"],
                "total": d["total"],
                "won": d["won"],
                "lost": d["lost"],
                "conversion_rate": round((d["won"] / (d["won"] + d["lost"]) * 100), 1) if (d["won"] + d["lost"]) > 0 else 0
            } for d in dealers if d["_id"]]
        }
    
    # Top Performers (for state and dealer)
    if entity_type in ["state", "dealer"]:
        if entity_type == "state":
            perf_field = "dealer"
        else:
            perf_field = "employee_name"
        
        top_pipeline = [
            {"$match": base_filter},
            {"$group": {
                "_id": f"${perf_field}",
                "total": {"$sum": 1},
                "won": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", won_stages]}, 1, 0]}},
                "lost": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", lost_stages]}, 1, 0]}}
            }},
            {"$match": {"$expr": {"$gt": [{"$add": ["$won", "$lost"]}, 5]}}},  # Min 5 closed leads
            {"$addFields": {
                "conversion_rate": {
                    "$cond": [
                        {"$gt": [{"$add": ["$won", "$lost"]}, 0]},
                        {"$multiply": [{"$divide": ["$won", {"$add": ["$won", "$lost"]}]}, 100]},
                        0
                    ]
                }
            }},
            {"$sort": {"conversion_rate": -1}},
            {"$limit": 3}
        ]
        top_performers = await db.leads.aggregate(top_pipeline).to_list(3)
        profile["top_performers"] = [{
            "name": t["_id"],
            "total": t["total"],
            "won": t["won"],
            "conversion_rate": round(t["conversion_rate"], 1)
        } for t in top_performers if t["_id"]]
    
    # Recent Leads (last 7 days, paginated)
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
    recent_filter = {**base_filter, "enquiry_date": {"$gte": seven_days_ago}}
    
    # Remove date range filter for recent leads query
    if "enquiry_date" in base_filter:
        recent_filter = {filter_field: entity_id, "enquiry_date": {"$gte": seven_days_ago}}
    
    recent_leads = await db.leads.find(
        recent_filter,
        {"_id": 0, "lead_id": 1, "enquiry_no": 1, "name": 1, "enquiry_date": 1, 
         "enquiry_stage": 1, "dealer": 1, "employee_name": 1, "segment": 1, "source": 1}
    ).sort("enquiry_date", -1).limit(10).to_list(10)
    
    recent_total = await db.leads.count_documents(recent_filter)
    
    profile["recent_leads"] = {
        "leads": recent_leads,
        "total": recent_total,
        "showing": len(recent_leads)
    }
    
    # Activity Timeline (recent activities for this entity's leads)
    lead_ids = await db.leads.distinct("lead_id", {filter_field: entity_id})
    activities = await db.activity_logs.find(
        {"resource": "lead", "resource_id": {"$in": lead_ids[:100]}},
        {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    profile["activity_timeline"] = activities
    
    return profile


@router.get("/recent-leads/{entity_type}/{entity_id}")
async def get_entity_recent_leads(
    request: Request,
    entity_type: str,
    entity_id: str,
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=50),
    current_user: User = Depends(get_current_user)
):
    """Get paginated recent leads for an entity (last 7 days)"""
    db = await get_db(request)
    
    field_map = {
        "state": "state",
        "dealer": "dealer",
        "city": "area",
        "employee": "employee_name"
    }
    
    if entity_type not in field_map:
        raise HTTPException(status_code=400, detail="Invalid entity type")
    
    filter_field = field_map[entity_type]
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
    
    query_filter = {
        filter_field: entity_id,
        "enquiry_date": {"$gte": seven_days_ago}
    }
    
    skip = (page - 1) * limit
    total = await db.leads.count_documents(query_filter)
    
    leads = await db.leads.find(
        query_filter,
        {"_id": 0, "lead_id": 1, "enquiry_no": 1, "name": 1, "enquiry_date": 1,
         "enquiry_stage": 1, "dealer": 1, "employee_name": 1, "segment": 1, 
         "source": 1, "state": 1, "area": 1, "phone_number": 1}
    ).sort("enquiry_date", -1).skip(skip).limit(limit).to_list(limit)
    
    return {
        "leads": leads,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit
    }


@router.get("/export/{entity_type}/{entity_id}")
async def export_entity_leads(
    request: Request,
    entity_type: str,
    entity_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export all leads for an entity to Excel"""
    from fastapi.responses import StreamingResponse
    import io
    
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    db = await get_db(request)
    
    field_map = {
        "state": "state",
        "dealer": "dealer",
        "city": "area",
        "employee": "employee_name"
    }
    
    if entity_type not in field_map:
        raise HTTPException(status_code=400, detail="Invalid entity type")
    
    filter_field = field_map[entity_type]
    query_filter = {filter_field: entity_id}
    
    if start_date and end_date:
        query_filter["enquiry_date"] = {"$gte": start_date, "$lte": end_date}
    
    leads = await db.leads.find(query_filter, {"_id": 0}).to_list(10000)
    
    if not leads:
        raise HTTPException(status_code=404, detail="No leads found")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{entity_type.capitalize()} Leads"
    
    # Headers
    if leads:
        headers = list(leads[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, lead in enumerate(leads, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=lead.get(header, ""))
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{entity_type}_{entity_id}_leads.xlsx".replace(" ", "_")
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# Entity Profile Configuration Endpoints
DEFAULT_ENTITY_PROFILE_CONFIG = {
    "config_id": "global",
    "kpis": {
        "enabled_kpis": [
            "total_leads", "open_leads", "won_leads", "lost_leads", 
            "conversion_rate", "avg_lead_age", "avg_closure_time"
        ],
        "show_call_quotation_kpis": True
    },
    "charts": {
        "stage_breakdown": {"enabled": True, "title": "Lead Stage Breakdown"},
        "source_breakdown": {"enabled": True, "title": "Lead Source Distribution"},
        "segment_performance": {"enabled": True, "title": "Segment Performance"},
        "trend": {"enabled": True, "title": "Lead Trend Over Time"},
        "followup_status": {"enabled": True, "title": "Follow-up Status"}
    },
    "sub_entities": {
        "show_top_performers": True,
        "show_employees": True,
        "show_dealers": True,
        "show_cities": True
    },
    "display_options": {
        "use_dashboard_date_filter": True,
        "default_trend_months": 6
    }
}


@router.get("/config")
async def get_entity_profile_config(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Get entity profile page configuration"""
    db = await get_db(request)
    
    config = await db.entity_profile_config.find_one({"config_id": "global"}, {"_id": 0})
    
    if not config:
        # Return default config
        return DEFAULT_ENTITY_PROFILE_CONFIG
    
    return config


@router.put("/config")
async def update_entity_profile_config(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Update entity profile page configuration (Admin only)"""
    if current_user.role not in ["Admin", "Manager"]:
        raise HTTPException(status_code=403, detail="Only Admin/Manager can update configuration")
    
    db = await get_db(request)
    data = await request.json()
    
    # Merge with existing config
    existing = await db.entity_profile_config.find_one({"config_id": "global"})
    
    config = {
        "config_id": "global",
        "kpis": data.get("kpis", DEFAULT_ENTITY_PROFILE_CONFIG["kpis"]),
        "charts": data.get("charts", DEFAULT_ENTITY_PROFILE_CONFIG["charts"]),
        "sub_entities": data.get("sub_entities", DEFAULT_ENTITY_PROFILE_CONFIG["sub_entities"]),
        "display_options": data.get("display_options", DEFAULT_ENTITY_PROFILE_CONFIG["display_options"]),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.name or current_user.email
    }
    
    if existing:
        await db.entity_profile_config.update_one(
            {"config_id": "global"},
            {"$set": config}
        )
    else:
        await db.entity_profile_config.insert_one(config)
    
    return {"message": "Entity profile configuration updated", "config": config}


@router.get("/available-kpis")
async def get_available_kpis(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Get list of available KPIs that can be shown on entity profile pages"""
    db = await get_db(request)
    
    # Get all active metrics from metric_settings
    metrics = await db.metric_settings.find(
        {"is_active": True},
        {"_id": 0, "metric_id": 1, "display_name": 1, "description": 1, "unit": 1, "color": 1, "icon": 1}
    ).to_list(50)
    
    # Add built-in metrics that are always available
    built_in_metrics = [
        {"metric_id": "total_leads", "display_name": "Total Leads", "unit": "", "is_built_in": True},
        {"metric_id": "conversion_rate", "display_name": "Conversion Rate", "unit": "%", "is_built_in": True},
        {"metric_id": "avg_lead_age", "display_name": "Avg Lead Age", "unit": "days", "is_built_in": True},
        {"metric_id": "avg_closure_time", "display_name": "Avg Closure Time", "unit": "days", "is_built_in": True},
        {"metric_id": "calls_placed", "display_name": "Calls Placed", "unit": "", "is_built_in": True},
        {"metric_id": "not_called", "display_name": "Not Called", "unit": "", "is_built_in": True},
        {"metric_id": "quotations_sent", "display_name": "Quotations Sent", "unit": "", "is_built_in": True},
        {"metric_id": "call_to_quotation_rate", "display_name": "Call to Quotation Rate", "unit": "%", "is_built_in": True},
    ]
    
    return {
        "built_in_metrics": built_in_metrics,
        "configurable_metrics": metrics
    }



# ============================================================================
# ENHANCED ENTITY ANALYTICS - Summary Builder, Market Share, KVA, YoY, etc.
# ============================================================================

def get_fy_for_date(date_str: str) -> tuple:
    """Get FY start and end dates for a given date string"""
    from datetime import datetime as dt
    date = dt.strptime(date_str[:10], "%Y-%m-%d")
    if date.month >= 4:
        fy_start = f"{date.year}-04-01"
        fy_end = f"{date.year + 1}-03-31"
    else:
        fy_start = f"{date.year - 1}-04-01"
        fy_end = f"{date.year}-03-31"
    return fy_start, fy_end


def get_last_fy_dates(start_date: str, end_date: str) -> tuple:
    """Calculate last FY dates for YoY comparison"""
    from datetime import datetime as dt
    start_dt = dt.strptime(start_date, "%Y-%m-%d")
    end_dt = dt.strptime(end_date, "%Y-%m-%d")
    ly_start = start_dt.replace(year=start_dt.year - 1).strftime("%Y-%m-%d")
    ly_end = end_dt.replace(year=end_dt.year - 1).strftime("%Y-%m-%d")
    return ly_start, ly_end


@router.get("/enhanced-analytics/{entity_type}/{entity_id}")
async def get_enhanced_entity_analytics(
    request: Request,
    entity_type: str,
    entity_id: str,
    current_user: User = Depends(get_current_user),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    time_frame: str = Query("monthly", enum=["monthly", "quarterly", "yearly"]),
    breakdown_by: str = Query("segment", enum=["segment", "employee", "dealer", "source", "kva", "district"])
):
    """
    Enhanced analytics for entity profiles including:
    - Mini Summary Builder (time-based breakdown)
    - Market Share calculations
    - KVA-wise breakdown (individual values)
    - YoY comparisons
    - Rank/Position
    - Pipeline health
    - Lead age distribution
    - Top segments
    """
    db = await get_db(request)
    
    field_map = {
        "state": "state",
        "dealer": "dealer", 
        "city": "area",
        "employee": "employee_name"
    }
    
    if entity_type not in field_map:
        raise HTTPException(status_code=400, detail="Invalid entity type")
    
    filter_field = field_map[entity_type]
    
    # Use FY dates if not provided
    if not start_date or not end_date:
        start_date, end_date = get_indian_fy_dates()
    
    # Calculate last year dates for YoY
    ly_start, ly_end = get_last_fy_dates(start_date, end_date)
    
    # Standardized won stages
    WON_STAGES = ["Closed-Won", "Order Booked"]
    LOST_STAGES = ["Closed-Lost", "Closed-Dropped"]
    
    # Base filter
    base_filter = {
        filter_field: entity_id,
        "is_deleted": {"$ne": True},
        "deleted_at": {"$exists": False},
        "enquiry_date": {"$gte": start_date, "$lte": end_date},
        "$or": [
            {"is_transferred": {"$exists": False}},
            {"is_transferred": False},
            {"is_transferred": None}
        ]
    }
    
    # Last year filter
    ly_filter = {
        filter_field: entity_id,
        "is_deleted": {"$ne": True},
        "deleted_at": {"$exists": False},
        "enquiry_date": {"$gte": ly_start, "$lte": ly_end},
        "$or": [
            {"is_transferred": {"$exists": False}},
            {"is_transferred": False},
            {"is_transferred": None}
        ]
    }
    
    result = {
        "entity_type": entity_type,
        "entity_id": entity_id,
        "date_range": {"start_date": start_date, "end_date": end_date},
        "ly_date_range": {"start_date": ly_start, "end_date": ly_end}
    }
    
    # ===== 1. SUMMARY BUILDER (Mini version) =====
    # Time frame extraction using FY logic
    time_frame_map = {
        "monthly": {"$substr": ["$enquiry_date", 0, 7]},
        "quarterly": {
            "$concat": [
                {"$cond": {
                    "if": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["01", "02", "03"]]},
                    "then": {"$toString": {"$subtract": [{"$toInt": {"$substr": ["$enquiry_date", 0, 4]}}, 1]}},
                    "else": {"$substr": ["$enquiry_date", 0, 4]}
                }},
                "-",
                {"$cond": {
                    "if": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["01", "02", "03"]]},
                    "then": {"$substr": ["$enquiry_date", 2, 2]},
                    "else": {"$toString": {"$mod": [{"$add": [{"$toInt": {"$substr": ["$enquiry_date", 0, 4]}}, 1]}, 100]}}
                }},
                "-Q",
                {"$switch": {
                    "branches": [
                        {"case": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["04", "05", "06"]]}, "then": "1"},
                        {"case": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["07", "08", "09"]]}, "then": "2"},
                        {"case": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["10", "11", "12"]]}, "then": "3"},
                        {"case": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["01", "02", "03"]]}, "then": "4"}
                    ],
                    "default": "1"
                }}
            ]
        },
        "yearly": {
            "$concat": [
                "FY",
                {"$cond": {
                    "if": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["01", "02", "03"]]},
                    "then": {"$toString": {"$subtract": [{"$toInt": {"$substr": ["$enquiry_date", 0, 4]}}, 1]}},
                    "else": {"$substr": ["$enquiry_date", 0, 4]}
                }},
                "-",
                {"$cond": {
                    "if": {"$in": [{"$substr": ["$enquiry_date", 5, 2]}, ["01", "02", "03"]]},
                    "then": {"$substr": ["$enquiry_date", 2, 2]},
                    "else": {"$toString": {"$mod": [{"$add": [{"$toInt": {"$substr": ["$enquiry_date", 0, 4]}}, 1]}, 100]}}
                }}
            ]
        }
    }
    
    time_extraction = time_frame_map.get(time_frame, time_frame_map["monthly"])
    
    summary_pipeline = [
        {"$match": base_filter},
        {"$addFields": {"time_period": time_extraction}},
        {"$group": {
            "_id": "$time_period",
            "total_leads": {"$sum": 1},
            "won_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
            "lost_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", LOST_STAGES]}, 1, 0]}},
            "open_leads": {"$sum": {"$cond": [{"$eq": ["$enquiry_status", "Open"]}, 1, 0]}},
            "total_kva": {"$sum": {"$ifNull": ["$kva", 0]}},
            "hot_leads": {"$sum": {"$cond": [{"$eq": ["$enquiry_type", "Hot"]}, 1, 0]}},
            "warm_leads": {"$sum": {"$cond": [{"$eq": ["$enquiry_type", "Warm"]}, 1, 0]}},
            "cold_leads": {"$sum": {"$cond": [{"$eq": ["$enquiry_type", "Cold"]}, 1, 0]}}
        }},
        {"$addFields": {
            "conversion_rate": {
                "$cond": [
                    {"$gt": [{"$add": ["$won_leads", "$lost_leads"]}, 0]},
                    {"$multiply": [{"$divide": ["$won_leads", {"$add": ["$won_leads", "$lost_leads"]}]}, 100]},
                    0
                ]
            }
        }},
        {"$sort": {"_id": 1}}
    ]
    
    summary_data = await db.leads.aggregate(summary_pipeline).to_list(100)
    result["summary_builder"] = [{
        "period": s["_id"],
        "total_leads": s["total_leads"],
        "won_leads": s["won_leads"],
        "lost_leads": s["lost_leads"],
        "open_leads": s["open_leads"],
        "total_kva": round(s["total_kva"], 1),
        "conversion_rate": round(s["conversion_rate"], 1),
        "hot_leads": s["hot_leads"],
        "warm_leads": s["warm_leads"],
        "cold_leads": s["cold_leads"]
    } for s in summary_data]
    
    # ===== 2. MARKET SHARE CALCULATIONS (using market potential) =====
    # Get entity's won leads (without has_so_record filter for accurate count)
    entity_won = await db.leads.count_documents({
        **base_filter,
        "enquiry_stage": {"$in": WON_STAGES}
    })
    
    # Get total market potential from market_potential_districts
    total_market_potential = 0
    async for doc in db.market_potential_districts.find():
        total_market_potential += doc.get("potential", 0)
    
    market_share = {
        "entity_wins": entity_won,
        "total_market_potential": total_market_potential,
        "share_of_company": round((entity_won / total_market_potential * 100) if total_market_potential > 0 else 0, 2)
    }
    
    # For dealers - get state and district market shares based on market potential
    if entity_type == "dealer":
        # Get dealer's primary state from leads (most common state)
        state_pipeline = [
            {"$match": {filter_field: entity_id, "is_deleted": {"$ne": True}}},
            {"$group": {"_id": "$state", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 1}
        ]
        state_result = await db.leads.aggregate(state_pipeline).to_list(1)
        dealer_state = state_result[0]["_id"] if state_result else None
        
        # Get dealer's allocated districts from market_potential_districts
        dealer_districts = await db.market_potential_districts.find(
            {"dealer": entity_id}
        ).to_list(100)
        
        if dealer_state:
            # Calculate state potential (sum of all districts in dealer's primary state)
            state_potential = 0
            async for doc in db.market_potential_districts.find({"state": dealer_state}):
                state_potential += doc.get("potential", 0)
            
            # Get dealer's won in that state
            entity_state_won = await db.leads.count_documents({
                **base_filter,
                "state": dealer_state,
                "enquiry_stage": {"$in": WON_STAGES}
            })
            
            market_share["state"] = dealer_state
            market_share["state_potential"] = state_potential
            market_share["state_wins"] = entity_state_won
            market_share["share_of_state"] = round((entity_state_won / state_potential * 100) if state_potential > 0 else 0, 2)
        
        # District-wise breakdown for dealer (all allocated districts)
        if dealer_districts:
            district_shares = []
            for dist_doc in dealer_districts:
                district_name = dist_doc.get("district")
                district_potential = dist_doc.get("potential", 0)
                
                # Get dealer's won in this district
                district_won = await db.leads.count_documents({
                    **base_filter,
                    "district": district_name,
                    "enquiry_stage": {"$in": WON_STAGES}
                })
                
                district_shares.append({
                    "district": district_name,
                    "won": district_won,
                    "potential": district_potential,
                    "share": round((district_won / district_potential * 100) if district_potential > 0 else 0, 2)
                })
            
            # Sort by share descending
            district_shares.sort(key=lambda x: x["share"], reverse=True)
            market_share["district_breakdown"] = district_shares
    
    # For employees - get dealer share
    elif entity_type == "employee":
        sample = await db.leads.find_one({filter_field: entity_id, "is_deleted": {"$ne": True}})
        if sample and sample.get("dealer"):
            dealer_name = sample["dealer"]
            # Get dealer's total potential
            dealer_potential = 0
            async for doc in db.market_potential_districts.find({"dealer": dealer_name}):
                dealer_potential += doc.get("potential", 0)
            
            # Get employee's won
            emp_won = await db.leads.count_documents({
                **base_filter,
                "enquiry_stage": {"$in": WON_STAGES}
            })
            
            market_share["dealer"] = dealer_name
            market_share["dealer_potential"] = dealer_potential
            market_share["share_of_dealer"] = round((emp_won / dealer_potential * 100) if dealer_potential > 0 else 0, 2)
    
    # For states - calculate state's share of total market
    elif entity_type == "state":
        state_potential = 0
        async for doc in db.market_potential_districts.find({"state": entity_id}):
            state_potential += doc.get("potential", 0)
        
        market_share["state_potential"] = state_potential
        market_share["share_of_market"] = round((entity_won / state_potential * 100) if state_potential > 0 else 0, 2)
    
    result["market_share"] = market_share
    
    # ===== 3. KVA-WISE BREAKDOWN (Individual values) =====
    kva_pipeline = [
        {"$match": {**base_filter, "kva": {"$exists": True, "$ne": None, "$gt": 0}}},
        {"$group": {
            "_id": "$kva",
            "total_leads": {"$sum": 1},
            "won_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
            "lost_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", LOST_STAGES]}, 1, 0]}},
            "open_leads": {"$sum": {"$cond": [{"$eq": ["$enquiry_status", "Open"]}, 1, 0]}}
        }},
        {"$addFields": {
            "conversion_rate": {
                "$cond": [
                    {"$gt": [{"$add": ["$won_leads", "$lost_leads"]}, 0]},
                    {"$multiply": [{"$divide": ["$won_leads", {"$add": ["$won_leads", "$lost_leads"]}]}, 100]},
                    0
                ]
            }
        }},
        {"$sort": {"_id": 1}}
    ]
    
    kva_data = await db.leads.aggregate(kva_pipeline).to_list(100)
    result["kva_breakdown"] = [{
        "kva": k["_id"],
        "total_leads": k["total_leads"],
        "won_leads": k["won_leads"],
        "lost_leads": k["lost_leads"],
        "open_leads": k["open_leads"],
        "conversion_rate": round(k["conversion_rate"], 1)
    } for k in kva_data]
    
    # ===== 4. YoY COMPARISON =====
    # Current year totals
    cy_total = await db.leads.count_documents(base_filter)
    cy_won = await db.leads.count_documents({**base_filter, "enquiry_stage": {"$in": WON_STAGES}})
    cy_lost = await db.leads.count_documents({**base_filter, "enquiry_stage": {"$in": LOST_STAGES}})
    cy_kva = await db.leads.aggregate([
        {"$match": {**base_filter, "enquiry_stage": {"$in": WON_STAGES}}},
        {"$group": {"_id": None, "total_kva": {"$sum": {"$ifNull": ["$kva", 0]}}}}
    ]).to_list(1)
    cy_total_kva = cy_kva[0]["total_kva"] if cy_kva else 0
    
    # Last year totals
    ly_total = await db.leads.count_documents(ly_filter)
    ly_won = await db.leads.count_documents({**ly_filter, "enquiry_stage": {"$in": WON_STAGES}})
    ly_lost = await db.leads.count_documents({**ly_filter, "enquiry_stage": {"$in": LOST_STAGES}})
    ly_kva = await db.leads.aggregate([
        {"$match": {**ly_filter, "enquiry_stage": {"$in": WON_STAGES}}},
        {"$group": {"_id": None, "total_kva": {"$sum": {"$ifNull": ["$kva", 0]}}}}
    ]).to_list(1)
    ly_total_kva = ly_kva[0]["total_kva"] if ly_kva else 0
    
    def calc_yoy_change(current, last):
        if last == 0:
            return 100 if current > 0 else 0
        return round(((current - last) / last) * 100, 1)
    
    result["yoy_comparison"] = {
        "current_year": {
            "total_leads": cy_total,
            "won_leads": cy_won,
            "lost_leads": cy_lost,
            "total_kva": round(cy_total_kva, 1),
            "conversion_rate": round((cy_won / (cy_won + cy_lost) * 100) if (cy_won + cy_lost) > 0 else 0, 1)
        },
        "last_year": {
            "total_leads": ly_total,
            "won_leads": ly_won,
            "lost_leads": ly_lost,
            "total_kva": round(ly_total_kva, 1),
            "conversion_rate": round((ly_won / (ly_won + ly_lost) * 100) if (ly_won + ly_lost) > 0 else 0, 1)
        },
        "yoy_change": {
            "total_leads": calc_yoy_change(cy_total, ly_total),
            "won_leads": calc_yoy_change(cy_won, ly_won),
            "total_kva": calc_yoy_change(cy_total_kva, ly_total_kva)
        }
    }
    
    # ===== 5. RANK/POSITION =====
    # Calculate rank among peers
    if entity_type == "dealer":
        # Rank among all dealers
        rank_pipeline = [
            {"$match": {
                "is_deleted": {"$ne": True},
                "deleted_at": {"$exists": False},
                "enquiry_date": {"$gte": start_date, "$lte": end_date},
                "enquiry_stage": {"$in": WON_STAGES}
            }},
            {"$group": {"_id": "$dealer", "wins": {"$sum": 1}}},
            {"$sort": {"wins": -1}}
        ]
        all_dealers = await db.leads.aggregate(rank_pipeline).to_list(1000)
        total_dealers = len(all_dealers)
        rank = next((i + 1 for i, d in enumerate(all_dealers) if d["_id"] == entity_id), total_dealers)
        result["rank"] = {
            "position": rank,
            "total": total_dealers,
            "percentile": round((1 - rank / total_dealers) * 100, 1) if total_dealers > 0 else 0
        }
    elif entity_type == "employee":
        # Rank among employees in same dealer
        sample = await db.leads.find_one({filter_field: entity_id, "is_deleted": {"$ne": True}})
        dealer = sample.get("dealer") if sample else None
        if dealer:
            rank_pipeline = [
                {"$match": {
                    "dealer": dealer,
                    "is_deleted": {"$ne": True},
                    "deleted_at": {"$exists": False},
                    "enquiry_date": {"$gte": start_date, "$lte": end_date},
                    "enquiry_stage": {"$in": WON_STAGES}
                }},
                {"$group": {"_id": "$employee_name", "wins": {"$sum": 1}}},
                {"$sort": {"wins": -1}}
            ]
            all_employees = await db.leads.aggregate(rank_pipeline).to_list(1000)
            total_employees = len(all_employees)
            rank = next((i + 1 for i, e in enumerate(all_employees) if e["_id"] == entity_id), total_employees)
            result["rank"] = {
                "position": rank,
                "total": total_employees,
                "within": dealer,
                "percentile": round((1 - rank / total_employees) * 100, 1) if total_employees > 0 else 0
            }
    
    # ===== 6. PIPELINE HEALTH (Hot/Warm/Cold distribution) =====
    pipeline_health = [
        {"$match": {**base_filter, "enquiry_status": "Open"}},
        {"$group": {
            "_id": "$enquiry_type",
            "count": {"$sum": 1},
            "total_kva": {"$sum": {"$ifNull": ["$kva", 0]}}
        }}
    ]
    health_data = await db.leads.aggregate(pipeline_health).to_list(10)
    health_map = {h["_id"]: {"count": h["count"], "kva": round(h["total_kva"], 1)} for h in health_data}
    
    total_open = sum(h["count"] for h in health_data)
    result["pipeline_health"] = {
        "hot": health_map.get("Hot", {"count": 0, "kva": 0}),
        "warm": health_map.get("Warm", {"count": 0, "kva": 0}),
        "cold": health_map.get("Cold", {"count": 0, "kva": 0}),
        "total_open": total_open,
        "distribution": {
            "hot_pct": round((health_map.get("Hot", {}).get("count", 0) / total_open * 100) if total_open > 0 else 0, 1),
            "warm_pct": round((health_map.get("Warm", {}).get("count", 0) / total_open * 100) if total_open > 0 else 0, 1),
            "cold_pct": round((health_map.get("Cold", {}).get("count", 0) / total_open * 100) if total_open > 0 else 0, 1)
        }
    }
    
    # ===== 7. LEAD AGE DISTRIBUTION =====
    age_pipeline = [
        {"$match": {**base_filter, "enquiry_status": "Open"}},
        {"$addFields": {
            "lead_age": {
                "$dateDiff": {
                    "startDate": {
                        "$cond": {
                            "if": {"$eq": [{"$type": "$enquiry_date"}, "date"]},
                            "then": "$enquiry_date",
                            "else": {"$dateFromString": {"dateString": "$enquiry_date", "onError": None}}
                        }
                    },
                    "endDate": "$$NOW",
                    "unit": "day"
                }
            }
        }},
        {"$match": {"lead_age": {"$ne": None}}},
        {"$bucket": {
            "groupBy": "$lead_age",
            "boundaries": [0, 31, 61, 91, 10000],
            "default": "90+",
            "output": {
                "count": {"$sum": 1},
                "total_kva": {"$sum": {"$ifNull": ["$kva", 0]}}
            }
        }}
    ]
    
    age_data = await db.leads.aggregate(age_pipeline).to_list(10)
    age_buckets = {
        "0-30": {"count": 0, "kva": 0},
        "31-60": {"count": 0, "kva": 0},
        "61-90": {"count": 0, "kva": 0},
        "90+": {"count": 0, "kva": 0}
    }
    
    for a in age_data:
        bucket_id = a["_id"]
        if bucket_id == 0:
            age_buckets["0-30"] = {"count": a["count"], "kva": round(a["total_kva"], 1)}
        elif bucket_id == 31:
            age_buckets["31-60"] = {"count": a["count"], "kva": round(a["total_kva"], 1)}
        elif bucket_id == 61:
            age_buckets["61-90"] = {"count": a["count"], "kva": round(a["total_kva"], 1)}
        elif bucket_id == 91 or bucket_id == "90+":
            age_buckets["90+"] = {"count": a["count"], "kva": round(a["total_kva"], 1)}
    
    result["lead_age_distribution"] = age_buckets
    
    # ===== 8. TOP SEGMENTS =====
    segment_pipeline = [
        {"$match": base_filter},
        {"$group": {
            "_id": "$segment",
            "total_leads": {"$sum": 1},
            "won_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
            "total_kva": {"$sum": {"$ifNull": ["$kva", 0]}}
        }},
        {"$addFields": {
            "conversion_rate": {
                "$cond": [
                    {"$gt": ["$total_leads", 0]},
                    {"$multiply": [{"$divide": ["$won_leads", "$total_leads"]}, 100]},
                    0
                ]
            }
        }},
        {"$sort": {"won_leads": -1}},
        {"$limit": 10}
    ]
    
    segment_data = await db.leads.aggregate(segment_pipeline).to_list(10)
    result["top_segments"] = [{
        "segment": s["_id"] or "Unknown",
        "total_leads": s["total_leads"],
        "won_leads": s["won_leads"],
        "total_kva": round(s["total_kva"], 1),
        "conversion_rate": round(s["conversion_rate"], 1)
    } for s in segment_data]
    
    # ===== 9. RECENT ACTIVITY (Last 10 leads) =====
    recent_leads = await db.leads.find(
        {filter_field: entity_id, "is_deleted": {"$ne": True}},
        {"_id": 0, "enquiry_no": 1, "name": 1, "enquiry_date": 1, "enquiry_stage": 1, 
         "enquiry_type": 1, "kva": 1, "segment": 1, "phone_number": 1}
    ).sort("enquiry_date", -1).limit(10).to_list(10)
    
    result["recent_activity"] = recent_leads
    
    # ===== 10. DIMENSION BREAKDOWN (based on breakdown_by param) =====
    breakdown_field_map = {
        "segment": "$segment",
        "employee": "$employee_name",
        "dealer": "$dealer",
        "source": "$source",
        "kva": "$kva",
        "district": "$district"
    }
    
    breakdown_field = breakdown_field_map.get(breakdown_by, "$segment")
    
    breakdown_pipeline = [
        {"$match": base_filter},
        {"$group": {
            "_id": breakdown_field,
            "total_leads": {"$sum": 1},
            "won_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
            "lost_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", LOST_STAGES]}, 1, 0]}},
            "open_leads": {"$sum": {"$cond": [{"$eq": ["$enquiry_status", "Open"]}, 1, 0]}},
            "total_kva": {"$sum": {"$ifNull": ["$kva", 0]}}
        }},
        {"$addFields": {
            "conversion_rate": {
                "$cond": [
                    {"$gt": [{"$add": ["$won_leads", "$lost_leads"]}, 0]},
                    {"$multiply": [{"$divide": ["$won_leads", {"$add": ["$won_leads", "$lost_leads"]}]}, 100]},
                    0
                ]
            }
        }},
        {"$sort": {"total_leads": -1}},
        {"$limit": 20}
    ]
    
    breakdown_data = await db.leads.aggregate(breakdown_pipeline).to_list(20)
    result["breakdown"] = {
        "by": breakdown_by,
        "data": [{
            "name": b["_id"] if b["_id"] else "Unknown",
            "total_leads": b["total_leads"],
            "won_leads": b["won_leads"],
            "lost_leads": b["lost_leads"],
            "open_leads": b["open_leads"],
            "total_kva": round(b["total_kva"], 1),
            "conversion_rate": round(b["conversion_rate"], 1)
        } for b in breakdown_data]
    }
    
    return result
