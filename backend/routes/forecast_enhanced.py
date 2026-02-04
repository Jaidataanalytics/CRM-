"""
Enhanced Forecast Module - Advanced Analytics and Export

Features:
- Dealer-wise KVA range forecast
- Dealer-wise District split forecast
- Auto-adjustment for consistency
- Seasonality analysis
- Lead-to-Win conversion time
- Product mix trends
- Geographic expansion opportunities
- Confidence intervals (3 scenarios + range bands)
- Excel export with charts as images
- Enhanced saved projections with audit trail
"""

from fastapi import APIRouter, HTTPException, Request, Depends, Response
from typing import Optional, List, Dict, Tuple
from datetime import datetime, timezone, timedelta
from dateutil.relativedelta import relativedelta
import logging
import io
import base64
import json
import uuid
from statistics import mean, stdev, median
from collections import defaultdict

from models.user import User, UserRole
from routes.auth import get_current_user, require_roles

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/forecast-enhanced", tags=["Enhanced Forecast"])

# Won stages
WON_STAGES = ["Closed-Won", "Order Booked"]
LOST_STAGES = ["Closed-Lost", "Closed-Dropped"]

# Month extraction pipeline for parsing "DD Mon YYYY" format dates
def get_month_extraction_pipeline():
    """Returns MongoDB aggregation expression to extract YYYY-MM from various date formats."""
    return {
        "$switch": {
            "branches": [
                # Format: "DD Mon YYYY" (e.g., "01 Aug 2025")
                {
                    "case": {
                        "$regexMatch": {
                            "input": "$enquiry_date",
                            "regex": "^\\d{2} [A-Za-z]{3} \\d{4}"
                        }
                    },
                    "then": {
                        "$concat": [
                            {"$substr": ["$enquiry_date", 7, 4]},
                            "-",
                            {"$switch": {
                                "branches": [
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Jan"]}, "then": "01"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Feb"]}, "then": "02"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Mar"]}, "then": "03"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Apr"]}, "then": "04"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "May"]}, "then": "05"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Jun"]}, "then": "06"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Jul"]}, "then": "07"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Aug"]}, "then": "08"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Sep"]}, "then": "09"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Oct"]}, "then": "10"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Nov"]}, "then": "11"},
                                    {"case": {"$eq": [{"$substr": ["$enquiry_date", 3, 3]}, "Dec"]}, "then": "12"},
                                ],
                                "default": "00"
                            }}
                        ]
                    }
                },
                # Format: "YYYY-MM-DD"
                {
                    "case": {
                        "$regexMatch": {
                            "input": "$enquiry_date",
                            "regex": "^\\d{4}-\\d{2}"
                        }
                    },
                    "then": {"$substr": ["$enquiry_date", 0, 7]}
                }
            ],
            "default": "unknown"
        }
    }


async def get_db(request: Request):
    return request.app.state.db


def get_last_12_months_range():
    """Get date range for last 12 months"""
    end_date = datetime.now(timezone.utc)
    start_date = end_date - relativedelta(months=12)
    return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")


# ============================================
# DEALER-KVA FORECAST
# ============================================

@router.get("/dealer-kva-forecast")
async def get_dealer_kva_forecast(
    request: Request,
    months_ahead: int = 3,
    current_user: User = Depends(get_current_user)
):
    """
    Generate dealer-wise KVA range forecast.
    Uses last 12 months of historical data to predict future sales by dealer and KVA.
    """
    db = await get_db(request)
    start_date, end_date = get_last_12_months_range()
    
    # Get historical data grouped by dealer, kva, and month
    pipeline = [
        {
            "$match": {
                "enquiry_date": {"$gte": start_date, "$lte": end_date},
                "enquiry_stage": {"$in": WON_STAGES},
                "deleted_at": {"$exists": False},
                "dealer": {"$exists": True, "$ne": None, "$ne": ""},
                "kva": {"$exists": True, "$ne": None}
            }
        },
        {
            "$addFields": {
                "month": get_month_extraction_pipeline(),
                "kva_value": {
                    "$cond": [
                        {"$isNumber": "$kva"},
                        "$kva",
                        {"$toDouble": {"$ifNull": ["$kva", 0]}}
                    ]
                }
            }
        },
        {
            "$group": {
                "_id": {
                    "dealer": "$dealer",
                    "kva": "$kva_value",
                    "month": "$month"
                },
                "units_sold": {"$sum": {"$ifNull": ["$qty", 1]}},
                "count": {"$sum": 1}
            }
        },
        {"$sort": {"_id.dealer": 1, "_id.kva": 1, "_id.month": 1}}
    ]
    
    results = await db.leads.aggregate(pipeline).to_list(5000)
    
    # Process results into dealer-kva structure
    dealer_kva_history = defaultdict(lambda: defaultdict(list))
    all_kvas = set()
    all_dealers = set()
    
    for r in results:
        dealer = r["_id"]["dealer"]
        kva = r["_id"]["kva"]
        month = r["_id"]["month"]
        units = r["units_sold"]
        
        if dealer and kva:
            dealer_kva_history[dealer][kva].append({
                "month": month,
                "units": units
            })
            all_kvas.add(kva)
            all_dealers.add(dealer)
    
    # Generate forecasts
    forecasts = []
    dealer_totals = {}
    
    for dealer in sorted(all_dealers):
        dealer_forecast = {
            "dealer": dealer,
            "kva_breakdown": [],
            "total_units": 0
        }
        
        for kva in sorted(all_kvas):
            history = dealer_kva_history[dealer].get(kva, [])
            
            if not history:
                continue
            
            # Calculate average monthly sales
            monthly_units = [h["units"] for h in history]
            avg_monthly = mean(monthly_units) if monthly_units else 0
            
            # Predict for next N months
            predicted_units = int(round(avg_monthly * months_ahead))
            
            if predicted_units > 0:
                dealer_forecast["kva_breakdown"].append({
                    "kva": kva,
                    "predicted_units": predicted_units,
                    "avg_monthly": round(avg_monthly, 1),
                    "historical_months": len(monthly_units)
                })
                dealer_forecast["total_units"] += predicted_units
        
        if dealer_forecast["kva_breakdown"]:
            forecasts.append(dealer_forecast)
            dealer_totals[dealer] = dealer_forecast["total_units"]
    
    # Sort by total units descending
    forecasts.sort(key=lambda x: x["total_units"], reverse=True)
    
    # Calculate grand totals
    grand_total = sum(f["total_units"] for f in forecasts)
    kva_totals = defaultdict(int)
    for f in forecasts:
        for kb in f["kva_breakdown"]:
            kva_totals[kb["kva"]] += kb["predicted_units"]
    
    return {
        "success": True,
        "forecast_horizon_months": months_ahead,
        "data_period": {"start": start_date, "end": end_date},
        "dealer_forecasts": forecasts,
        "kva_summary": [
            {"kva": k, "total_units": v}
            for k, v in sorted(kva_totals.items())
        ],
        "grand_total_units": grand_total,
        "dealer_count": len(forecasts)
    }


# ============================================
# DEALER-DISTRICT FORECAST
# ============================================

@router.get("/dealer-district-forecast")
async def get_dealer_district_forecast(
    request: Request,
    months_ahead: int = 3,
    current_user: User = Depends(get_current_user)
):
    """
    Generate dealer-wise district split forecast.
    Uses last 12 months of historical data to predict future sales by dealer and district.
    ONLY shows districts where each dealer has actually made sales (historical presence).
    """
    db = await get_db(request)
    start_date, end_date = get_last_12_months_range()
    
    # Get historical data grouped by dealer, district, and month
    # This automatically filters to only dealer-district combinations that exist
    pipeline = [
        {
            "$match": {
                "enquiry_date": {"$gte": start_date, "$lte": end_date},
                "enquiry_stage": {"$in": WON_STAGES},
                "deleted_at": {"$exists": False},
                "dealer": {"$exists": True, "$ne": None, "$ne": ""},
                "district": {"$exists": True, "$ne": None, "$ne": ""}
            }
        },
        {
            "$addFields": {
                "month": get_month_extraction_pipeline()
            }
        },
        {
            "$group": {
                "_id": {
                    "dealer": "$dealer",
                    "district": "$district",
                    "month": "$month"
                },
                "units_sold": {"$sum": {"$ifNull": ["$qty", 1]}},
                "count": {"$sum": 1}
            }
        },
        {"$sort": {"_id.dealer": 1, "_id.district": 1, "_id.month": 1}}
    ]
    
    results = await db.leads.aggregate(pipeline).to_list(5000)
    
    # Process results - only include dealer-district pairs that have historical data
    dealer_district_history = defaultdict(lambda: defaultdict(list))
    
    for r in results:
        dealer = r["_id"]["dealer"]
        district = r["_id"]["district"]
        month = r["_id"]["month"]
        units = r["units_sold"]
        
        if dealer and district:
            dealer_district_history[dealer][district].append({
                "month": month,
                "units": units
            })
    
    # Generate forecasts - each dealer only gets districts they've sold in
    forecasts = []
    
    for dealer in sorted(dealer_district_history.keys()):
        dealer_districts = dealer_district_history[dealer]
        
        dealer_forecast = {
            "dealer": dealer,
            "district_breakdown": [],
            "total_units": 0,
            "districts_count": 0
        }
        
        for district, history in dealer_districts.items():
            if not history:
                continue
            
            # Calculate average monthly sales
            monthly_units = [h["units"] for h in history]
            avg_monthly = mean(monthly_units) if monthly_units else 0
            
            # Predict for next N months
            predicted_units = int(round(avg_monthly * months_ahead))
            
            if predicted_units > 0:
                dealer_forecast["district_breakdown"].append({
                    "district": district,
                    "predicted_units": predicted_units,
                    "avg_monthly": round(avg_monthly, 1),
                    "historical_months": len(monthly_units)
                })
                dealer_forecast["total_units"] += predicted_units
        
        # Sort districts by predicted units
        dealer_forecast["district_breakdown"].sort(key=lambda x: x["predicted_units"], reverse=True)
        dealer_forecast["districts_count"] = len(dealer_forecast["district_breakdown"])
        
        if dealer_forecast["district_breakdown"]:
            forecasts.append(dealer_forecast)
    
    # Sort by total units descending
    forecasts.sort(key=lambda x: x["total_units"], reverse=True)
    
    # Calculate district totals (across all dealers)
    district_totals = defaultdict(int)
    for f in forecasts:
        for db_item in f["district_breakdown"]:
            district_totals[db_item["district"]] += db_item["predicted_units"]
    
    grand_total = sum(f["total_units"] for f in forecasts)
    
    return {
        "success": True,
        "forecast_horizon_months": months_ahead,
        "data_period": {"start": start_date, "end": end_date},
        "dealer_forecasts": forecasts,
        "district_summary": [
            {"district": k, "total_units": v}
            for k, v in sorted(district_totals.items(), key=lambda x: x[1], reverse=True)
        ],
        "grand_total_units": grand_total,
        "dealer_count": len(forecasts),
        "note": "Each dealer only shows districts where they have historical sales"
    }


# ============================================
# SEASONALITY ANALYSIS
# ============================================

@router.get("/seasonality-analysis")
async def get_seasonality_analysis(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Analyze seasonality patterns - which months historically perform best.
    """
    db = await get_db(request)
    
    # Get all historical data grouped by month
    pipeline = [
        {
            "$match": {
                "enquiry_date": {"$exists": True, "$ne": None},
                "deleted_at": {"$exists": False}
            }
        },
        {
            "$addFields": {
                "month_num": {"$month": {"$dateFromString": {"dateString": "$enquiry_date", "format": "%Y-%m-%d", "onError": None}}},
                "year": {"$year": {"$dateFromString": {"dateString": "$enquiry_date", "format": "%Y-%m-%d", "onError": None}}}
            }
        },
        {
            "$match": {"month_num": {"$ne": None}}
        },
        {
            "$group": {
                "_id": {"month": "$month_num", "year": "$year"},
                "total_leads": {"$sum": 1},
                "won_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
                "total_qty": {"$sum": {"$ifNull": ["$qty", 0]}}
            }
        },
        {"$sort": {"_id.month": 1, "_id.year": 1}}
    ]
    
    results = await db.leads.aggregate(pipeline).to_list(500)
    
    # Process by month
    month_data = defaultdict(lambda: {"leads": [], "won": [], "qty": []})
    
    for r in results:
        month = r["_id"]["month"]
        month_data[month]["leads"].append(r["total_leads"])
        month_data[month]["won"].append(r["won_leads"])
        month_data[month]["qty"].append(r["total_qty"])
    
    month_names = ["", "January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    
    seasonality = []
    for month in range(1, 13):
        data = month_data[month]
        if data["leads"]:
            seasonality.append({
                "month": month,
                "month_name": month_names[month],
                "avg_leads": round(mean(data["leads"]), 1),
                "avg_won": round(mean(data["won"]), 1),
                "avg_qty": round(mean(data["qty"]), 1),
                "years_of_data": len(data["leads"]),
                "min_leads": min(data["leads"]),
                "max_leads": max(data["leads"])
            })
    
    # Calculate overall average for comparison
    all_leads = [s["avg_leads"] for s in seasonality if s["avg_leads"] > 0]
    overall_avg = mean(all_leads) if all_leads else 0
    
    # Add seasonality index (100 = average)
    for s in seasonality:
        if overall_avg > 0:
            s["seasonality_index"] = round((s["avg_leads"] / overall_avg) * 100, 1)
        else:
            s["seasonality_index"] = 100
    
    # Find best and worst months
    best_month = max(seasonality, key=lambda x: x["avg_leads"]) if seasonality else None
    worst_month = min(seasonality, key=lambda x: x["avg_leads"]) if seasonality else None
    
    return {
        "success": True,
        "seasonality": seasonality,
        "best_month": best_month,
        "worst_month": worst_month,
        "overall_avg_leads": round(overall_avg, 1),
        "insights": [
            f"Best performing month: {best_month['month_name']} (index: {best_month['seasonality_index']})" if best_month else "",
            f"Weakest month: {worst_month['month_name']} (index: {worst_month['seasonality_index']})" if worst_month else ""
        ]
    }


# ============================================
# LEAD-TO-WIN CONVERSION TIME
# ============================================

@router.get("/conversion-time-analysis")
async def get_conversion_time_analysis(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Analyze lead-to-win conversion time by dealer and segment.
    """
    db = await get_db(request)
    
    # Get won leads with enquiry and closure dates
    pipeline = [
        {
            "$match": {
                "enquiry_stage": {"$in": WON_STAGES},
                "enquiry_date": {"$exists": True, "$ne": None},
                "deleted_at": {"$exists": False}
            }
        },
        {
            "$addFields": {
                "closure_date_parsed": {
                    "$cond": [
                        {"$and": [
                            {"$ne": ["$enquiry_closure_date", None]},
                            {"$ne": ["$enquiry_closure_date", ""]}
                        ]},
                        "$enquiry_closure_date",
                        "$updated_at"
                    ]
                }
            }
        },
        {
            "$project": {
                "dealer": 1,
                "segment": 1,
                "enquiry_date": 1,
                "closure_date": "$closure_date_parsed",
                "kva": 1
            }
        }
    ]
    
    results = await db.leads.aggregate(pipeline).to_list(10000)
    
    # Calculate conversion days
    dealer_times = defaultdict(list)
    segment_times = defaultdict(list)
    overall_times = []
    
    for r in results:
        try:
            enquiry_date = datetime.strptime(r["enquiry_date"][:10], "%Y-%m-%d")
            closure_str = r.get("closure_date", "")
            if closure_str:
                closure_date = datetime.strptime(closure_str[:10], "%Y-%m-%d")
                days = (closure_date - enquiry_date).days
                
                if 0 <= days <= 365:  # Reasonable range
                    overall_times.append(days)
                    
                    dealer = r.get("dealer", "Unknown")
                    if dealer:
                        dealer_times[dealer].append(days)
                    
                    segment = r.get("segment", "Unknown")
                    if segment:
                        segment_times[segment].append(days)
        except:
            continue
    
    # Calculate statistics
    def calc_stats(times_list):
        if not times_list:
            return {"avg": 0, "median": 0, "min": 0, "max": 0, "count": 0}
        return {
            "avg": round(mean(times_list), 1),
            "median": round(median(times_list), 1),
            "min": min(times_list),
            "max": max(times_list),
            "count": len(times_list)
        }
    
    dealer_analysis = [
        {"dealer": d, **calc_stats(times)}
        for d, times in dealer_times.items()
        if len(times) >= 5  # Minimum sample size
    ]
    dealer_analysis.sort(key=lambda x: x["avg"])
    
    segment_analysis = [
        {"segment": s, **calc_stats(times)}
        for s, times in segment_times.items()
        if len(times) >= 5
    ]
    segment_analysis.sort(key=lambda x: x["avg"])
    
    return {
        "success": True,
        "overall_stats": calc_stats(overall_times),
        "by_dealer": dealer_analysis[:20],  # Top 20
        "by_segment": segment_analysis,
        "insights": [
            f"Average conversion time: {calc_stats(overall_times)['avg']} days",
            f"Fastest dealer: {dealer_analysis[0]['dealer']} ({dealer_analysis[0]['avg']} days avg)" if dealer_analysis else "",
            f"Slowest dealer: {dealer_analysis[-1]['dealer']} ({dealer_analysis[-1]['avg']} days avg)" if dealer_analysis else ""
        ]
    }


# ============================================
# PRODUCT MIX TRENDS
# ============================================

@router.get("/product-mix-trends")
async def get_product_mix_trends(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Analyze which KVA products are growing or declining.
    """
    db = await get_db(request)
    
    # Get last 24 months of data
    end_date = datetime.now(timezone.utc)
    start_date = end_date - relativedelta(months=24)
    
    pipeline = [
        {
            "$match": {
                "enquiry_date": {"$gte": start_date.strftime("%Y-%m-%d")},
                "enquiry_stage": {"$in": WON_STAGES},
                "deleted_at": {"$exists": False},
                "kva": {"$exists": True, "$ne": None, "$ne": ""}
            }
        },
        {
            "$addFields": {
                "year_half": {"$substr": ["$enquiry_date", 0, 7]}  # Just use YYYY-MM
            }
        },
        {
            "$group": {
                "_id": {"kva": "$kva", "period": "$year_half"},
                "units": {"$sum": {"$ifNull": ["$qty", 1]}},
                "count": {"$sum": 1}
            }
        },
        {"$sort": {"_id.kva": 1, "_id.period": 1}}
    ]
    
    results = await db.leads.aggregate(pipeline).to_list(1000)
    
    # Process by KVA
    kva_trends = defaultdict(lambda: {"periods": {}})
    periods = set()
    
    for r in results:
        kva = r["_id"]["kva"]
        period = r["_id"]["period"]
        kva_trends[kva]["periods"][period] = r["units"]
        periods.add(period)
    
    sorted_periods = sorted(periods)
    
    # Calculate trends
    product_analysis = []
    for kva, data in kva_trends.items():
        period_values = [data["periods"].get(p, 0) for p in sorted_periods]
        
        if len(period_values) >= 2 and period_values[0] > 0:
            first_half = mean(period_values[:len(period_values)//2]) if period_values[:len(period_values)//2] else 0
            second_half = mean(period_values[len(period_values)//2:]) if period_values[len(period_values)//2:] else 0
            
            if first_half > 0:
                growth_rate = ((second_half - first_half) / first_half) * 100
            else:
                growth_rate = 100 if second_half > 0 else 0
            
            trend = "growing" if growth_rate > 10 else "declining" if growth_rate < -10 else "stable"
            
            product_analysis.append({
                "kva": kva,
                "total_units": sum(period_values),
                "growth_rate": round(growth_rate, 1),
                "trend": trend,
                "period_breakdown": {p: data["periods"].get(p, 0) for p in sorted_periods}
            })
    
    product_analysis.sort(key=lambda x: x["growth_rate"], reverse=True)
    
    growing = [p for p in product_analysis if p["trend"] == "growing"]
    declining = [p for p in product_analysis if p["trend"] == "declining"]
    
    return {
        "success": True,
        "periods": sorted_periods,
        "products": product_analysis,
        "growing_products": growing[:5],
        "declining_products": declining[:5],
        "insights": [
            f"Top growing KVA: {growing[0]['kva']} (+{growing[0]['growth_rate']}%)" if growing else "No growing products",
            f"Most declining KVA: {declining[-1]['kva']} ({declining[-1]['growth_rate']}%)" if declining else "No declining products"
        ]
    }


# ============================================
# GEOGRAPHIC EXPANSION OPPORTUNITIES
# ============================================

@router.get("/geographic-opportunities")
async def get_geographic_opportunities(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Identify districts with low penetration but high potential.
    """
    db = await get_db(request)
    start_date, end_date = get_last_12_months_range()
    
    # Get lead and won counts by district
    pipeline = [
        {
            "$match": {
                "enquiry_date": {"$gte": start_date},
                "deleted_at": {"$exists": False},
                "district": {"$exists": True, "$ne": None, "$ne": ""}
            }
        },
        {
            "$group": {
                "_id": {"district": "$district", "state": "$state"},
                "total_leads": {"$sum": 1},
                "won_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
                "total_qty": {"$sum": {"$ifNull": ["$qty", 0]}}
            }
        },
        {"$sort": {"total_leads": -1}}
    ]
    
    results = await db.leads.aggregate(pipeline).to_list(500)
    
    # Calculate conversion rates and identify opportunities
    districts = []
    for r in results:
        total = r["total_leads"]
        won = r["won_leads"]
        conv_rate = (won / total * 100) if total > 0 else 0
        
        districts.append({
            "district": r["_id"]["district"],
            "state": r["_id"].get("state") or "Unknown",
            "total_leads": total,
            "won_leads": won,
            "conversion_rate": round(conv_rate, 1),
            "total_qty": r["total_qty"]
        })
    
    # Calculate averages
    avg_leads = mean([d["total_leads"] for d in districts]) if districts else 0
    avg_conv = mean([d["conversion_rate"] for d in districts]) if districts else 0
    
    # Identify opportunities: high conversion but low volume
    opportunities = [
        d for d in districts
        if d["conversion_rate"] > avg_conv and d["total_leads"] < avg_leads
    ]
    opportunities.sort(key=lambda x: x["conversion_rate"], reverse=True)
    
    # Identify underperformers: high volume but low conversion
    underperformers = [
        d for d in districts
        if d["conversion_rate"] < avg_conv and d["total_leads"] > avg_leads
    ]
    underperformers.sort(key=lambda x: x["total_leads"], reverse=True)
    
    return {
        "success": True,
        "all_districts": districts[:50],
        "expansion_opportunities": opportunities[:10],
        "underperforming_districts": underperformers[:10],
        "averages": {
            "avg_leads_per_district": round(avg_leads, 1),
            "avg_conversion_rate": round(avg_conv, 1)
        },
        "insights": [
            f"Found {len(opportunities)} high-potential expansion districts",
            f"Found {len(underperformers)} underperforming districts needing attention"
        ]
    }


# ============================================
# CONFIDENCE INTERVALS / SCENARIOS
# ============================================

@router.get("/forecast-scenarios")
async def get_forecast_scenarios(
    request: Request,
    months_ahead: int = 3,
    current_user: User = Depends(get_current_user)
):
    """
    Generate forecast with 3 scenarios: pessimistic, realistic, optimistic.
    Also provides confidence bands (±15%, ±25%).
    """
    db = await get_db(request)
    start_date, end_date = get_last_12_months_range()
    
    # Get monthly historical data
    pipeline = [
        {
            "$match": {
                "enquiry_date": {"$gte": start_date, "$lte": end_date},
                "deleted_at": {"$exists": False}
            }
        },
        {
            "$addFields": {
                "month": get_month_extraction_pipeline()
            }
        },
        {
            "$group": {
                "_id": "$month",
                "total_leads": {"$sum": 1},
                "won_leads": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
                "total_qty": {"$sum": {"$ifNull": ["$qty", 0]}}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    results = await db.leads.aggregate(pipeline).to_list(100)
    
    if len(results) < 6:
        raise HTTPException(status_code=400, detail="Not enough historical data for scenario forecasting")
    
    # Calculate statistics
    leads_values = [r["total_leads"] for r in results]
    won_values = [r["won_leads"] for r in results]
    qty_values = [r["total_qty"] for r in results]
    
    leads_avg = mean(leads_values)
    leads_std = stdev(leads_values) if len(leads_values) > 1 else leads_avg * 0.2
    
    won_avg = mean(won_values)
    won_std = stdev(won_values) if len(won_values) > 1 else won_avg * 0.2
    
    # Generate scenarios
    scenarios = {
        "pessimistic": {
            "factor": 0.85,
            "description": "Conservative estimate assuming market challenges"
        },
        "realistic": {
            "factor": 1.0,
            "description": "Most likely outcome based on historical trends"
        },
        "optimistic": {
            "factor": 1.15,
            "description": "Best case scenario with favorable conditions"
        }
    }
    
    # Generate monthly predictions
    last_month = results[-1]["_id"]
    year, month = int(last_month[:4]), int(last_month[5:7])
    
    predictions = []
    for i in range(months_ahead):
        month += 1
        if month > 12:
            month = 1
            year += 1
        
        period = f"{year}-{str(month).zfill(2)}"
        
        base_leads = leads_avg
        base_won = won_avg
        
        period_pred = {
            "period": period,
            "scenarios": {}
        }
        
        for scenario_name, scenario_config in scenarios.items():
            factor = scenario_config["factor"]
            period_pred["scenarios"][scenario_name] = {
                "leads": int(round(base_leads * factor)),
                "won": int(round(base_won * factor)),
                "description": scenario_config["description"]
            }
        
        # Confidence bands
        period_pred["confidence_bands"] = {
            "leads": {
                "point_estimate": int(round(base_leads)),
                "low_15": int(round(base_leads * 0.85)),
                "high_15": int(round(base_leads * 1.15)),
                "low_25": int(round(base_leads * 0.75)),
                "high_25": int(round(base_leads * 1.25))
            },
            "won": {
                "point_estimate": int(round(base_won)),
                "low_15": int(round(base_won * 0.85)),
                "high_15": int(round(base_won * 1.15)),
                "low_25": int(round(base_won * 0.75)),
                "high_25": int(round(base_won * 1.25))
            }
        }
        
        predictions.append(period_pred)
    
    # Calculate totals for each scenario
    scenario_totals = {}
    for scenario_name in scenarios.keys():
        scenario_totals[scenario_name] = {
            "total_leads": sum(p["scenarios"][scenario_name]["leads"] for p in predictions),
            "total_won": sum(p["scenarios"][scenario_name]["won"] for p in predictions)
        }
    
    return {
        "success": True,
        "months_ahead": months_ahead,
        "predictions": predictions,
        "scenario_totals": scenario_totals,
        "historical_stats": {
            "avg_monthly_leads": round(leads_avg, 1),
            "std_monthly_leads": round(leads_std, 1),
            "avg_monthly_won": round(won_avg, 1),
            "data_months": len(results)
        }
    }


# ============================================
# MONTHLY DETAILED FORECAST (NEW)
# ============================================

@router.get("/monthly-forecast")
async def get_monthly_detailed_forecast(
    request: Request,
    months_ahead: int = 3,
    include_current_month: bool = True,
    current_user: User = Depends(get_current_user)
):
    """
    Generate detailed monthly forecast with leads, closures, and conversion rate.
    
    - include_current_month: If True, excludes current month's actual data and predicts it
    - Returns separate predictions for each month
    - Shows: total leads, closures (won), conversion rate, dealer breakdown
    """
    db = await get_db(request)
    
    # Calculate date ranges
    now = datetime.now(timezone.utc)
    current_month_str = now.strftime("%Y-%m")
    
    # Historical data: last 12 months, excluding current month if include_current_month is True
    end_date = now - relativedelta(months=1) if include_current_month else now
    start_date = end_date - relativedelta(months=12)
    
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")
    
    # Get monthly aggregated data: total leads and closures per month
    pipeline = [
        {
            "$match": {
                "enquiry_date": {"$gte": start_str, "$lte": end_str},
                "deleted_at": {"$exists": False},
            }
        },
        {
            "$addFields": {
                "month": get_month_extraction_pipeline(),
            }
        },
        {
            "$group": {
                "_id": "$month",
                "total_leads": {"$sum": 1},
                "closures": {
                    "$sum": {
                        "$cond": [
                            {"$in": ["$enquiry_stage", WON_STAGES]},
                            1,
                            0
                        ]
                    }
                },
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    monthly_history = await db.leads.aggregate(pipeline).to_list(100)
    
    # Also get dealer breakdown for closures
    dealer_pipeline = [
        {
            "$match": {
                "enquiry_date": {"$gte": start_str, "$lte": end_str},
                "enquiry_stage": {"$in": WON_STAGES},
                "deleted_at": {"$exists": False},
                "dealer": {"$exists": True, "$ne": None, "$ne": ""},
            }
        },
        {
            "$addFields": {
                "month": get_month_extraction_pipeline(),
            }
        },
        {
            "$group": {
                "_id": {
                    "dealer": "$dealer",
                    "month": "$month"
                },
                "closures": {"$sum": 1},
                "kva_sum": {"$sum": {"$ifNull": ["$kva", 0]}},
            }
        },
    ]
    
    dealer_history = await db.leads.aggregate(dealer_pipeline).to_list(10000)
    
    # Build dealer history structure
    dealer_monthly = defaultdict(lambda: defaultdict(lambda: {"closures": 0, "kva": 0}))
    all_dealers = set()
    for r in dealer_history:
        dealer = r["_id"]["dealer"]
        month = r["_id"]["month"]
        dealer_monthly[dealer][month]["closures"] = r["closures"]
        dealer_monthly[dealer][month]["kva"] = r["kva_sum"]
        all_dealers.add(dealer)
    
    # Calculate historical averages
    if not monthly_history:
        return {"success": False, "message": "No historical data available"}
    
    # Calculate overall averages
    total_months = len(monthly_history)
    avg_leads = sum(m["total_leads"] for m in monthly_history) / total_months
    avg_closures = sum(m["closures"] for m in monthly_history) / total_months
    avg_conversion = (avg_closures / avg_leads * 100) if avg_leads > 0 else 0
    
    # Calculate seasonality by month number
    month_leads = defaultdict(list)
    month_closures = defaultdict(list)
    for m in monthly_history:
        month_num = int(m["_id"].split("-")[1])
        month_leads[month_num].append(m["total_leads"])
        month_closures[month_num].append(m["closures"])
    
    # Generate future months to predict
    future_months = []
    start_month = now if include_current_month else now + relativedelta(months=1)
    for i in range(months_ahead):
        m = start_month + relativedelta(months=i)
        future_months.append(m.strftime("%Y-%m"))
    
    # Generate predictions for each month
    monthly_forecasts = []
    
    for future_month in future_months:
        month_num = int(future_month.split("-")[1])
        
        # Get seasonality factors for this month
        hist_leads = month_leads.get(month_num, [avg_leads])
        hist_closures = month_closures.get(month_num, [avg_closures])
        
        # Predict based on historical data for same month
        predicted_leads = int(round(mean(hist_leads) if hist_leads else avg_leads))
        predicted_closures = int(round(mean(hist_closures) if hist_closures else avg_closures))
        
        # Calculate conversion rate
        conversion_rate = round((predicted_closures / predicted_leads * 100), 1) if predicted_leads > 0 else 0
        
        # Calculate dealer breakdown (based on historical share)
        dealer_breakdown = []
        total_dealer_closures = sum(
            sum(dealer_monthly[d][m]["closures"] for m in dealer_monthly[d])
            for d in all_dealers
        )
        
        for dealer in sorted(all_dealers):
            dealer_total = sum(dealer_monthly[dealer][m]["closures"] for m in dealer_monthly[dealer])
            if dealer_total > 0:
                share = dealer_total / total_dealer_closures if total_dealer_closures > 0 else 0
                dealer_predicted = int(round(predicted_closures * share))
                if dealer_predicted > 0:
                    dealer_breakdown.append({
                        "dealer": dealer,
                        "predicted_closures": dealer_predicted,
                        "historical_closures": dealer_total,
                        "share_percentage": round(share * 100, 1)
                    })
        
        # Sort by predicted closures
        dealer_breakdown.sort(key=lambda x: x["predicted_closures"], reverse=True)
        
        month_forecast = {
            "month": future_month,
            "month_name": datetime.strptime(future_month, "%Y-%m").strftime("%B %Y"),
            "is_current_month": future_month == current_month_str,
            "predicted_leads": predicted_leads,
            "predicted_closures": predicted_closures,
            "conversion_rate": conversion_rate,
            "dealer_breakdown": dealer_breakdown[:10],  # Top 10 dealers
            "historical_same_month": {
                "leads": hist_leads,
                "closures": hist_closures
            }
        }
        
        monthly_forecasts.append(month_forecast)
    
    # Calculate grand totals
    grand_total_leads = sum(mf["predicted_leads"] for mf in monthly_forecasts)
    grand_total_closures = sum(mf["predicted_closures"] for mf in monthly_forecasts)
    
    # Build chart data for monthly bars
    chart_data = {
        "months": [mf["month_name"] for mf in monthly_forecasts],
        "leads": [mf["predicted_leads"] for mf in monthly_forecasts],
        "closures": [mf["predicted_closures"] for mf in monthly_forecasts],
        "conversion_rates": [mf["conversion_rate"] for mf in monthly_forecasts]
    }
    
    return {
        "success": True,
        "include_current_month": include_current_month,
        "current_month": current_month_str,
        "historical_period": {"start": start_str, "end": end_str, "months_analyzed": total_months},
        "forecast_months": future_months,
        "monthly_forecasts": monthly_forecasts,
        "grand_totals": {
            "leads": grand_total_leads,
            "closures": grand_total_closures,
            "avg_conversion": round((grand_total_closures / grand_total_leads * 100), 1) if grand_total_leads > 0 else 0
        },
        "historical_averages": {
            "avg_leads_per_month": round(avg_leads, 1),
            "avg_closures_per_month": round(avg_closures, 1),
            "avg_conversion_rate": round(avg_conversion, 1)
        },
        "chart_data": chart_data,
        "summary": {
            "total_dealers": len(all_dealers),
            "months_predicted": len(future_months)
        }
    }


# ============================================
# COMPREHENSIVE AI FORECAST WITH MODEL TESTING
# ============================================

@router.post("/comprehensive-forecast")
async def generate_comprehensive_forecast(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """
    Generate comprehensive AI-based forecast with:
    - Multiple ML model testing (ARIMA, XGBoost, RandomForest, etc.)
    - Model selection based on R², MAPE, MAE, RMSE
    - Organization-level predictions (total leads, closures)
    - Dealer-level breakdown with KVA and district splits
    - Consistency validation (all totals match)
    
    Two-level forecast:
    - Level 1: Organization (total, by dealer, by KVA, by district)
    - Level 2: Per dealer (by KVA, by district)
    
    Request body parameters:
    - months_ahead: int (default 3) - Number of months to forecast
    - include_current_month: bool (default True) - Include current month in predictions
    - years_back: int (default 3) - Years of historical data to use for training
    - force_model: str (optional) - Force a specific model instead of auto-selection
      Options: "Simple Moving Average", "Weighted Moving Average", "Exponential Smoothing",
               "Seasonal (Same-Month)", "Linear Trend", "ARIMA", "Random Forest",
               "Gradient Boosting", "XGBoost", "Ensemble (Hybrid)"
    """
    from routes.forecast_models import (
        ModelOptimizer, 
        SimpleMovingAverage, 
        WeightedMovingAverage,
        ExponentialSmoothing,
        SeasonalNaive,
        LinearTrend,
        ARIMAForecaster,
        RandomForestForecaster,
        GradientBoostingForecaster,
        XGBoostForecaster,
        EnsembleForecaster
    )
    
    db = await get_db(request)
    body = await request.json()
    
    months_ahead = body.get("months_ahead", 3)
    include_current_month = body.get("include_current_month", True)
    years_back = body.get("years_back", 3)  # User-configurable historical range
    force_model = body.get("force_model")  # Optional: force a specific model
    
    # Calculate date range based on years_back parameter
    now = datetime.now(timezone.utc)
    current_month = now.strftime("%Y-%m")
    
    # Calculate start date based on years_back (Indian FY starts April 1)
    start_year = now.year - years_back
    start_month = 4  # April
    start_date = f"{start_year}-{start_month:02d}-01"
    start_month_str = f"{start_year}-{start_month:02d}"
    
    end_date = (now - relativedelta(months=1)).strftime("%Y-%m-%d") if include_current_month else now.strftime("%Y-%m-%d")
    
    # ===== STEP 1: Get historical monthly data =====
    monthly_pipeline = [
        {
            "$match": {
                "enquiry_date": {"$exists": True, "$ne": None, "$ne": ""},
                "deleted_at": {"$exists": False}
            }
        },
        {"$addFields": {"month": get_month_extraction_pipeline()}},
        {"$match": {"month": {"$ne": "unknown", "$gte": start_month_str}}},
        {
            "$group": {
                "_id": "$month",
                "leads": {"$sum": 1},
                "closures": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}},
                "total_kva": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, {"$ifNull": ["$kva", 0]}, 0]}}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    
    monthly_data = await db.leads.aggregate(monthly_pipeline).to_list(100)
    
    if len(monthly_data) < 12:
        return {"success": False, "message": f"Need at least 12 months of data. Found {len(monthly_data)}. Try increasing years_back parameter."}
    
    # ===== STEP 2: Get dealer/KVA/district breakdown =====
    breakdown_pipeline = [
        {
            "$match": {
                "enquiry_date": {"$exists": True, "$ne": None, "$ne": ""},
                "deleted_at": {"$exists": False}
            }
        },
        {"$addFields": {"month": get_month_extraction_pipeline()}},
        {"$match": {"month": {"$ne": "unknown", "$gte": start_month_str}}},
        {
            "$group": {
                "_id": {
                    "month": "$month",
                    "dealer": {"$ifNull": ["$dealer", "Unknown"]},
                    "kva": {"$ifNull": ["$kva", 0]},
                    "district": {"$ifNull": ["$district", "Unknown"]}
                },
                "leads": {"$sum": 1},
                "closures": {"$sum": {"$cond": [{"$in": ["$enquiry_stage", WON_STAGES]}, 1, 0]}}
            }
        }
    ]
    
    breakdown_data = await db.leads.aggregate(breakdown_pipeline).to_list(100000)
    
    # Build historical structures
    dealer_history = defaultdict(lambda: {"leads": 0, "closures": 0, "months": set()})
    kva_history = defaultdict(lambda: {"leads": 0, "closures": 0})
    district_history = defaultdict(lambda: {"leads": 0, "closures": 0})
    dealer_kva_history = defaultdict(lambda: defaultdict(lambda: {"leads": 0, "closures": 0}))
    dealer_district_history = defaultdict(lambda: defaultdict(lambda: {"leads": 0, "closures": 0}))
    
    for r in breakdown_data:
        dealer = r["_id"]["dealer"]
        kva = r["_id"]["kva"]
        district = r["_id"]["district"]
        month = r["_id"]["month"]
        leads = r["leads"]
        closures = r["closures"]
        
        dealer_history[dealer]["leads"] += leads
        dealer_history[dealer]["closures"] += closures
        dealer_history[dealer]["months"].add(month)
        
        kva_history[kva]["leads"] += leads
        kva_history[kva]["closures"] += closures
        
        district_history[district]["leads"] += leads
        district_history[district]["closures"] += closures
        
        dealer_kva_history[dealer][kva]["leads"] += leads
        dealer_kva_history[dealer][kva]["closures"] += closures
        
        dealer_district_history[dealer][district]["leads"] += leads
        dealer_district_history[dealer][district]["closures"] += closures
    
    # ===== STEP 3: Run ML Model Selection =====
    # Prepare data for model optimizer
    leads_series = [{"month": d["_id"], "value": d["leads"]} for d in monthly_data]
    closures_series = [{"month": d["_id"], "value": d["closures"]} for d in monthly_data]
    
    # Model mapping for forced model selection
    MODEL_MAP = {
        "Simple Moving Average": SimpleMovingAverage,
        "Weighted Moving Average": WeightedMovingAverage,
        "Exponential Smoothing": ExponentialSmoothing,
        "Seasonal (Same-Month)": SeasonalNaive,
        "Linear Trend": LinearTrend,
        "ARIMA": ARIMAForecaster,
        "Random Forest": RandomForestForecaster,
        "Gradient Boosting": GradientBoostingForecaster,
        "XGBoost": XGBoostForecaster,
        "Ensemble (Hybrid)": EnsembleForecaster
    }
    
    if force_model and force_model in MODEL_MAP:
        # Use forced model instead of auto-selection
        ForcedModel = MODEL_MAP[force_model]
        
        # Initialize and fit leads model
        forced_leads_model = ForcedModel()
        forced_leads_model.fit([d["value"] for d in leads_series])
        leads_predictions = forced_leads_model.forecast(months_ahead)
        
        # Initialize and fit closures model
        forced_closures_model = ForcedModel()
        forced_closures_model.fit([d["value"] for d in closures_series])
        closures_predictions = forced_closures_model.forecast(months_ahead)
        
        # Create dummy optimizer results for metrics
        leads_optimizer = type('obj', (object,), {
            'best_model': forced_leads_model,
            'best_model_name': force_model,
            'best_accuracy': None,
            'best_mape': None,
            'results': [{"model": force_model, "accuracy": None, "mape": None, "note": "User selected"}]
        })()
        closures_optimizer = type('obj', (object,), {
            'best_model': forced_closures_model,
            'best_model_name': force_model,
            'best_accuracy': None,
            'best_mape': None,
            'results': [{"model": force_model, "accuracy": None, "mape": None, "note": "User selected"}]
        })()
        
        model_selection_mode = "manual"
    else:
        # Auto-select best model
        leads_optimizer = ModelOptimizer(leads_series)
        leads_result = leads_optimizer.optimize()
        
        closures_optimizer = ModelOptimizer(closures_series)
        closures_result = closures_optimizer.optimize()
        
        # Get predictions from best models
        leads_predictions = leads_optimizer.best_model.forecast(months_ahead) if leads_optimizer.best_model else []
        closures_predictions = closures_optimizer.best_model.forecast(months_ahead) if closures_optimizer.best_model else []
        
        model_selection_mode = "auto"
    
    # ===== STEP 4: Generate future months =====
    future_months = []
    start_month_dt = now if include_current_month else now + relativedelta(months=1)
    for i in range(months_ahead):
        m = start_month_dt + relativedelta(months=i)
        future_months.append(m.strftime("%Y-%m"))
    
    # Calculate historical totals for proportional distribution
    total_historical_leads = sum(d["leads"] for d in monthly_data)
    total_historical_closures = sum(d["closures"] for d in monthly_data)
    
    # ===== STEP 5: Build Organization-level forecast =====
    org_forecast = {
        "months": [],
        "totals": {
            "leads": 0,
            "closures": 0
        },
        "by_dealer": {},
        "by_kva": {},
        "by_district": {}
    }
    
    # Calculate dealer/KVA/district shares for proportional distribution
    dealer_lead_share = {d: h["leads"]/total_historical_leads for d, h in dealer_history.items() if total_historical_leads > 0}
    dealer_closure_share = {d: h["closures"]/total_historical_closures for d, h in dealer_history.items() if total_historical_closures > 0}
    
    kva_lead_share = {k: h["leads"]/total_historical_leads for k, h in kva_history.items() if total_historical_leads > 0}
    kva_closure_share = {k: h["closures"]/total_historical_closures for k, h in kva_history.items() if total_historical_closures > 0}
    
    district_lead_share = {d: h["leads"]/total_historical_leads for d, h in district_history.items() if total_historical_leads > 0}
    district_closure_share = {d: h["closures"]/total_historical_closures for d, h in district_history.items() if total_historical_closures > 0}
    
    # ===== STEP 6: Generate monthly predictions with breakdowns =====
    for i, month in enumerate(future_months):
        pred_leads = int(round(leads_predictions[i])) if i < len(leads_predictions) else 0
        pred_closures = int(round(closures_predictions[i])) if i < len(closures_predictions) else 0
        
        # Ensure closures don't exceed leads
        pred_closures = min(pred_closures, pred_leads)
        
        conversion_rate = round((pred_closures / pred_leads * 100), 1) if pred_leads > 0 else 0
        
        month_data = {
            "month": month,
            "month_name": datetime.strptime(month, "%Y-%m").strftime("%B %Y"),
            "is_current_month": month == current_month,
            "predicted_leads": pred_leads,
            "predicted_closures": pred_closures,
            "conversion_rate": conversion_rate,
            "dealer_breakdown": [],
            "kva_breakdown": [],
            "district_breakdown": []
        }
        
        # Distribute to dealers proportionally
        dealer_leads_sum = 0
        dealer_closures_sum = 0
        for dealer, share in sorted(dealer_lead_share.items(), key=lambda x: x[1], reverse=True):
            d_leads = int(round(pred_leads * share))
            d_closures = int(round(pred_closures * dealer_closure_share.get(dealer, share)))
            d_closures = min(d_closures, d_leads)  # Can't close more than leads
            
            if d_leads > 0 or d_closures > 0:
                dealer_leads_sum += d_leads
                dealer_closures_sum += d_closures
                month_data["dealer_breakdown"].append({
                    "dealer": dealer,
                    "predicted_leads": d_leads,
                    "predicted_closures": d_closures,
                    "conversion_rate": round((d_closures/d_leads*100), 1) if d_leads > 0 else 0,
                    "share_pct": round(share * 100, 2)
                })
        
        # Adjust for rounding errors
        if dealer_leads_sum != pred_leads and month_data["dealer_breakdown"]:
            month_data["dealer_breakdown"][0]["predicted_leads"] += (pred_leads - dealer_leads_sum)
        if dealer_closures_sum != pred_closures and month_data["dealer_breakdown"]:
            month_data["dealer_breakdown"][0]["predicted_closures"] += (pred_closures - dealer_closures_sum)
        
        # Distribute to KVA proportionally
        kva_leads_sum = 0
        kva_closures_sum = 0
        for kva, share in sorted(kva_lead_share.items(), key=lambda x: x[1], reverse=True):
            k_leads = int(round(pred_leads * share))
            k_closures = int(round(pred_closures * kva_closure_share.get(kva, share)))
            k_closures = min(k_closures, k_leads)
            
            if k_leads > 0 or k_closures > 0:
                kva_leads_sum += k_leads
                kva_closures_sum += k_closures
                month_data["kva_breakdown"].append({
                    "kva": kva,
                    "predicted_leads": k_leads,
                    "predicted_closures": k_closures,
                    "share_pct": round(share * 100, 2)
                })
        
        # Adjust for rounding errors
        if kva_leads_sum != pred_leads and month_data["kva_breakdown"]:
            month_data["kva_breakdown"][0]["predicted_leads"] += (pred_leads - kva_leads_sum)
        if kva_closures_sum != pred_closures and month_data["kva_breakdown"]:
            month_data["kva_breakdown"][0]["predicted_closures"] += (pred_closures - kva_closures_sum)
        
        # Distribute to districts proportionally
        dist_leads_sum = 0
        dist_closures_sum = 0
        for district, share in sorted(district_lead_share.items(), key=lambda x: x[1], reverse=True):
            di_leads = int(round(pred_leads * share))
            di_closures = int(round(pred_closures * district_closure_share.get(district, share)))
            di_closures = min(di_closures, di_leads)
            
            if di_leads > 0 or di_closures > 0:
                dist_leads_sum += di_leads
                dist_closures_sum += di_closures
                month_data["district_breakdown"].append({
                    "district": district,
                    "predicted_leads": di_leads,
                    "predicted_closures": di_closures,
                    "share_pct": round(share * 100, 2)
                })
        
        # Adjust for rounding errors
        if dist_leads_sum != pred_leads and month_data["district_breakdown"]:
            month_data["district_breakdown"][0]["predicted_leads"] += (pred_leads - dist_leads_sum)
        if dist_closures_sum != pred_closures and month_data["district_breakdown"]:
            month_data["district_breakdown"][0]["predicted_closures"] += (pred_closures - dist_closures_sum)
        
        org_forecast["months"].append(month_data)
        org_forecast["totals"]["leads"] += pred_leads
        org_forecast["totals"]["closures"] += pred_closures
    
    # ===== STEP 7: Build Dealer-level forecasts =====
    dealer_forecasts = {}
    
    for dealer in dealer_history.keys():
        if dealer == "Unknown":
            continue
            
        dealer_total_leads = dealer_history[dealer]["leads"]
        dealer_total_closures = dealer_history[dealer]["closures"]
        
        # Calculate KVA shares for this dealer
        dealer_kva_shares = {}
        for kva, data in dealer_kva_history[dealer].items():
            if dealer_total_leads > 0:
                dealer_kva_shares[kva] = {
                    "lead_share": data["leads"] / dealer_total_leads,
                    "closure_share": data["closures"] / dealer_total_closures if dealer_total_closures > 0 else 0
                }
        
        # Calculate district shares for this dealer
        dealer_district_shares = {}
        for district, data in dealer_district_history[dealer].items():
            if dealer_total_leads > 0:
                dealer_district_shares[district] = {
                    "lead_share": data["leads"] / dealer_total_leads,
                    "closure_share": data["closures"] / dealer_total_closures if dealer_total_closures > 0 else 0
                }
        
        dealer_forecast = {
            "dealer": dealer,
            "historical": {
                "total_leads": dealer_total_leads,
                "total_closures": dealer_total_closures,
                "months_active": len(dealer_history[dealer]["months"]),
                "conversion_rate": round((dealer_total_closures/dealer_total_leads*100), 1) if dealer_total_leads > 0 else 0
            },
            "months": []
        }
        
        # Get this dealer's predictions from org forecast
        for month_data in org_forecast["months"]:
            dealer_month = next(
                (d for d in month_data["dealer_breakdown"] if d["dealer"] == dealer),
                {"predicted_leads": 0, "predicted_closures": 0}
            )
            
            d_pred_leads = dealer_month["predicted_leads"]
            d_pred_closures = dealer_month["predicted_closures"]
            
            dm = {
                "month": month_data["month"],
                "month_name": month_data["month_name"],
                "predicted_leads": d_pred_leads,
                "predicted_closures": d_pred_closures,
                "conversion_rate": round((d_pred_closures/d_pred_leads*100), 1) if d_pred_leads > 0 else 0,
                "by_kva": [],
                "by_district": []
            }
            
            # Distribute to KVA for this dealer
            kva_sum_l, kva_sum_c = 0, 0
            for kva, shares in sorted(dealer_kva_shares.items(), key=lambda x: x[1]["lead_share"], reverse=True):
                k_leads = int(round(d_pred_leads * shares["lead_share"]))
                k_closures = int(round(d_pred_closures * shares["closure_share"]))
                k_closures = min(k_closures, k_leads)
                if k_leads > 0 or k_closures > 0:
                    kva_sum_l += k_leads
                    kva_sum_c += k_closures
                    dm["by_kva"].append({"kva": kva, "leads": k_leads, "closures": k_closures})
            
            # Adjust rounding
            if dm["by_kva"] and kva_sum_l != d_pred_leads:
                dm["by_kva"][0]["leads"] += (d_pred_leads - kva_sum_l)
            if dm["by_kva"] and kva_sum_c != d_pred_closures:
                dm["by_kva"][0]["closures"] += (d_pred_closures - kva_sum_c)
            
            # Distribute to districts for this dealer
            dist_sum_l, dist_sum_c = 0, 0
            for district, shares in sorted(dealer_district_shares.items(), key=lambda x: x[1]["lead_share"], reverse=True):
                di_leads = int(round(d_pred_leads * shares["lead_share"]))
                di_closures = int(round(d_pred_closures * shares["closure_share"]))
                di_closures = min(di_closures, di_leads)
                if di_leads > 0 or di_closures > 0:
                    dist_sum_l += di_leads
                    dist_sum_c += di_closures
                    dm["by_district"].append({"district": district, "leads": di_leads, "closures": di_closures})
            
            # Adjust rounding
            if dm["by_district"] and dist_sum_l != d_pred_leads:
                dm["by_district"][0]["leads"] += (d_pred_leads - dist_sum_l)
            if dm["by_district"] and dist_sum_c != d_pred_closures:
                dm["by_district"][0]["closures"] += (d_pred_closures - dist_sum_c)
            
            dealer_forecast["months"].append(dm)
        
        dealer_forecasts[dealer] = dealer_forecast
    
    # ===== STEP 8: Consistency validation =====
    consistency_check = {"passed": True, "issues": []}
    
    for month_data in org_forecast["months"]:
        month = month_data["month"]
        
        # Check dealer totals match
        dealer_lead_total = sum(d["predicted_leads"] for d in month_data["dealer_breakdown"])
        dealer_closure_total = sum(d["predicted_closures"] for d in month_data["dealer_breakdown"])
        
        if dealer_lead_total != month_data["predicted_leads"]:
            consistency_check["passed"] = False
            consistency_check["issues"].append(f"{month}: Dealer leads total ({dealer_lead_total}) != org total ({month_data['predicted_leads']})")
        
        # Check KVA totals match
        kva_lead_total = sum(k["predicted_leads"] for k in month_data["kva_breakdown"])
        kva_closure_total = sum(k["predicted_closures"] for k in month_data["kva_breakdown"])
        
        if kva_lead_total != month_data["predicted_leads"]:
            consistency_check["passed"] = False
            consistency_check["issues"].append(f"{month}: KVA leads total ({kva_lead_total}) != org total ({month_data['predicted_leads']})")
        
        # Check district totals match
        dist_lead_total = sum(d["predicted_leads"] for d in month_data["district_breakdown"])
        
        if dist_lead_total != month_data["predicted_leads"]:
            consistency_check["passed"] = False
            consistency_check["issues"].append(f"{month}: District leads total ({dist_lead_total}) != org total ({month_data['predicted_leads']})")
    
    # ===== STEP 9: Build model metrics =====
    model_metrics = {
        "model_selection_mode": model_selection_mode,
        "forced_model": force_model if model_selection_mode == "manual" else None,
        "available_models": list(MODEL_MAP.keys()),
        "leads_model": {
            "name": leads_optimizer.best_model.name if hasattr(leads_optimizer.best_model, 'name') else (force_model or "None"),
            "accuracy": round(leads_optimizer.best_accuracy, 2) if leads_optimizer.best_accuracy else None,
            "mape": round(leads_optimizer.best_mape, 2) if hasattr(leads_optimizer, 'best_mape') and leads_optimizer.best_mape else None,
            "all_models_tested": [
                {
                    "model": r["model"],
                    "accuracy": round(r.get("accuracy", 0), 2) if r.get("accuracy") else None,
                    "mape": round(r.get("mape", 100), 2) if r.get("mape") else None,
                    "note": r.get("note", "")
                }
                for r in leads_optimizer.results
            ]
        },
        "closures_model": {
            "name": closures_optimizer.best_model.name if hasattr(closures_optimizer.best_model, 'name') else (force_model or "None"),
            "accuracy": round(closures_optimizer.best_accuracy, 2) if closures_optimizer.best_accuracy else None,
            "mape": round(closures_optimizer.best_mape, 2) if hasattr(closures_optimizer, 'best_mape') and closures_optimizer.best_mape else None,
            "all_models_tested": [
                {
                    "model": r["model"],
                    "accuracy": round(r.get("accuracy", 0), 2) if r.get("accuracy") else None,
                    "mape": round(r.get("mape", 100), 2) if r.get("mape") else None,
                    "note": r.get("note", "")
                }
                for r in closures_optimizer.results
            ]
        },
        "data_quality": {
            "months_analyzed": len(monthly_data),
            "total_leads_in_training": total_historical_leads,
            "total_closures_in_training": total_historical_closures,
            "unique_dealers": len(dealer_history),
            "unique_kvas": len(kva_history),
            "unique_districts": len(district_history)
        }
    }
    
    # ===== STEP 10: Calculate summary statistics =====
    historical_avg_leads = total_historical_leads / len(monthly_data)
    historical_avg_closures = total_historical_closures / len(monthly_data)
    historical_conversion = (total_historical_closures / total_historical_leads * 100) if total_historical_leads > 0 else 0
    
    return {
        "success": True,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "forecast_horizon": months_ahead,
        "include_current_month": include_current_month,
        "years_back": years_back,
        "model_selection_mode": model_selection_mode,
        "forced_model": force_model if model_selection_mode == "manual" else None,
        "model_metrics": model_metrics,
        "consistency_check": consistency_check,
        "historical_summary": {
            "period": f"{start_date} to {end_date}",
            "months_analyzed": len(monthly_data),
            "avg_leads_per_month": round(historical_avg_leads, 1),
            "avg_closures_per_month": round(historical_avg_closures, 1),
            "avg_conversion_rate": round(historical_conversion, 1)
        },
        "organization_forecast": org_forecast,
        "dealer_forecasts": dealer_forecasts,
        "chart_data": {
            "months": [m["month_name"] for m in org_forecast["months"]],
            "leads": [m["predicted_leads"] for m in org_forecast["months"]],
            "closures": [m["predicted_closures"] for m in org_forecast["months"]],
            "conversion_rates": [m["conversion_rate"] for m in org_forecast["months"]]
        }
    }


# ============================================
# COMPREHENSIVE FORECAST EXPORT
# ============================================

@router.post("/export-excel")
async def export_forecast_to_excel(
    request: Request,
    current_user: User = Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER))
):
    """
    Export comprehensive forecast data to Excel with multiple sheets and embedded charts.
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    db = await get_db(request)
    body = await request.json()
    
    forecast_data = body.get("forecast_data", {})
    include_charts = body.get("include_charts", True)
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_data = [
        ["Forecast Export Summary"],
        [""],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Generated By", current_user.name or current_user.email],
        ["Forecast Horizon", f"{forecast_data.get('horizon_months', 3)} months"],
        [""],
        ["Key Metrics"],
    ]
    
    if "summary" in forecast_data:
        summary_data.append(["Summary", forecast_data["summary"]])
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Sheet 2: Monthly Forecast
    ws_monthly = wb.create_sheet("Monthly Forecast")
    predictions = forecast_data.get("predictions", [])
    
    if predictions:
        headers = ["Month", "Predicted Enquiries", "Predicted Closures", "Predicted KVA"]
        ws_monthly.append(headers)
        
        for pred in predictions:
            ws_monthly.append([
                pred.get("month", ""),
                pred.get("predicted_enquiries", 0),
                pred.get("predicted_closures", 0),
                pred.get("predicted_kva", 0)
            ])
    
    # Sheet 3: Dealer Forecast (if available)
    dealer_kva = forecast_data.get("dealer_kva_forecast", {})
    if dealer_kva and dealer_kva.get("dealer_forecasts"):
        ws_dealer = wb.create_sheet("Dealer-KVA Forecast")
        ws_dealer.append(["Dealer", "KVA", "Predicted Units", "Avg Monthly"])
        
        for dealer_data in dealer_kva["dealer_forecasts"]:
            dealer = dealer_data["dealer"]
            for kva_item in dealer_data.get("kva_breakdown", []):
                ws_dealer.append([
                    dealer,
                    kva_item["kva"],
                    kva_item["predicted_units"],
                    kva_item["avg_monthly"]
                ])
    
    # Sheet 4: District Forecast (if available)
    dealer_district = forecast_data.get("dealer_district_forecast", {})
    if dealer_district and dealer_district.get("dealer_forecasts"):
        ws_district = wb.create_sheet("Dealer-District Forecast")
        ws_district.append(["Dealer", "District", "Predicted Units", "Avg Monthly"])
        
        for dealer_data in dealer_district["dealer_forecasts"]:
            dealer = dealer_data["dealer"]
            for dist_item in dealer_data.get("district_breakdown", []):
                ws_district.append([
                    dealer,
                    dist_item["district"],
                    dist_item["predicted_units"],
                    dist_item["avg_monthly"]
                ])
    
    # Sheet 5: Scenarios
    scenarios = forecast_data.get("scenarios", {})
    if scenarios and scenarios.get("predictions"):
        ws_scenarios = wb.create_sheet("Forecast Scenarios")
        ws_scenarios.append(["Period", "Pessimistic Leads", "Realistic Leads", "Optimistic Leads",
                            "Pessimistic Won", "Realistic Won", "Optimistic Won"])
        
        for pred in scenarios["predictions"]:
            ws_scenarios.append([
                pred["period"],
                pred["scenarios"]["pessimistic"]["leads"],
                pred["scenarios"]["realistic"]["leads"],
                pred["scenarios"]["optimistic"]["leads"],
                pred["scenarios"]["pessimistic"]["won"],
                pred["scenarios"]["realistic"]["won"],
                pred["scenarios"]["optimistic"]["won"]
            ])
    
    # Sheet 6: Recommendations
    ws_reco = wb.create_sheet("Recommendations")
    ws_reco.append(["Recommendations and Insights"])
    ws_reco.append([""])
    
    recommendations = forecast_data.get("recommendations", [])
    if isinstance(recommendations, list):
        for reco in recommendations:
            if isinstance(reco, dict):
                ws_reco.append([reco.get("title", ""), reco.get("description", "")])
            else:
                ws_reco.append([str(reco)])
    
    # Add charts if requested
    if include_charts and predictions:
        # Create a simple line chart for monthly forecast
        chart = LineChart()
        chart.title = "Monthly Forecast"
        chart.style = 13
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Month"
        
        # Reference data for chart
        data = Reference(ws_monthly, min_col=2, min_row=1, max_col=3, max_row=len(predictions) + 1)
        cats = Reference(ws_monthly, min_col=1, min_row=2, max_row=len(predictions) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        ws_monthly.add_chart(chart, "F2")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"forecast_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# ============================================
# ENHANCED SAVE PROJECTION WITH AUDIT TRAIL
# ============================================

@router.post("/save-enhanced")
async def save_enhanced_forecast(
    request: Request,
    current_user: User = Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER))
):
    """
    Save forecast with all data, charts (as base64 images), and audit trail support.
    """
    db = await get_db(request)
    body = await request.json()
    
    forecast_data = body.get("forecast_data", {})
    chart_images = body.get("chart_images", {})  # Base64 encoded images
    notes = body.get("notes", "")
    
    projection_id = f"proj_{uuid.uuid4().hex[:12]}"
    
    projection = {
        "projection_id": projection_id,
        "saved_at": datetime.now(timezone.utc).isoformat(),
        "saved_by": {
            "user_id": current_user.user_id,
            "name": current_user.name or current_user.email,
            "email": current_user.email
        },
        "version": 1,
        "is_latest": True,
        
        # Core forecast data
        "forecast_data": forecast_data,
        
        # Chart images (base64)
        "chart_images": chart_images,
        
        # Additional analytics
        "dealer_kva_forecast": body.get("dealer_kva_forecast"),
        "dealer_district_forecast": body.get("dealer_district_forecast"),
        "scenarios": body.get("scenarios"),
        "seasonality": body.get("seasonality"),
        "conversion_analysis": body.get("conversion_analysis"),
        "product_trends": body.get("product_trends"),
        "geographic_opportunities": body.get("geographic_opportunities"),
        
        # Recommendations
        "recommendations": body.get("recommendations", []),
        "trends": body.get("trends", []),
        
        # User notes
        "notes": notes,
        
        # Audit trail
        "audit_trail": [{
            "action": "created",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "user": current_user.name or current_user.email,
            "details": "Initial projection created"
        }]
    }
    
    await db.enhanced_forecasts.insert_one(projection)
    
    return {
        "success": True,
        "message": "Enhanced forecast saved successfully",
        "projection_id": projection_id
    }


@router.put("/update-projection/{projection_id}")
async def update_projection(
    request: Request,
    projection_id: str,
    current_user: User = Depends(require_roles(UserRole.ADMIN, UserRole.MANAGER))
):
    """
    Update a saved projection with audit trail tracking.
    """
    db = await get_db(request)
    body = await request.json()
    
    # Find existing projection
    existing = await db.enhanced_forecasts.find_one({"projection_id": projection_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Projection not found")
    
    # Get changes and reason
    changes = body.get("changes", {})
    reason = body.get("reason", "Manual adjustment")
    
    # Create audit entry
    audit_entry = {
        "action": "updated",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "user": current_user.name or current_user.email,
        "details": reason,
        "changes": {
            "before": {},
            "after": {}
        }
    }
    
    # Track changes
    update_fields = {}
    for field, new_value in changes.items():
        if field in existing:
            audit_entry["changes"]["before"][field] = existing.get(field)
            audit_entry["changes"]["after"][field] = new_value
            update_fields[field] = new_value
    
    # Update version
    new_version = existing.get("version", 1) + 1
    update_fields["version"] = new_version
    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Push audit entry
    await db.enhanced_forecasts.update_one(
        {"projection_id": projection_id},
        {
            "$set": update_fields,
            "$push": {"audit_trail": audit_entry}
        }
    )
    
    return {
        "success": True,
        "message": "Projection updated successfully",
        "new_version": new_version
    }


@router.get("/projection/{projection_id}")
async def get_projection(
    request: Request,
    projection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a saved projection by ID with full data."""
    db = await get_db(request)
    
    projection = await db.enhanced_forecasts.find_one(
        {"projection_id": projection_id},
        {"_id": 0}
    )
    
    if not projection:
        raise HTTPException(status_code=404, detail="Projection not found")
    
    return {
        "success": True,
        "projection": projection
    }


@router.get("/projections")
async def list_projections(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """List all saved projections."""
    db = await get_db(request)
    
    cursor = db.enhanced_forecasts.find(
        {},
        {
            "_id": 0,
            "projection_id": 1,
            "saved_at": 1,
            "saved_by": 1,
            "version": 1,
            "notes": 1,
            "forecast_data.horizon_months": 1,
            "forecast_data.summary": 1
        }
    ).sort("saved_at", -1)
    
    projections = await cursor.to_list(100)
    
    return {
        "success": True,
        "projections": projections,
        "total": len(projections)
    }


@router.get("/projection/{projection_id}/audit-trail")
async def get_audit_trail(
    request: Request,
    projection_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get the audit trail for a projection."""
    db = await get_db(request)
    
    projection = await db.enhanced_forecasts.find_one(
        {"projection_id": projection_id},
        {"_id": 0, "audit_trail": 1, "version": 1}
    )
    
    if not projection:
        raise HTTPException(status_code=404, detail="Projection not found")
    
    return {
        "success": True,
        "projection_id": projection_id,
        "current_version": projection.get("version", 1),
        "audit_trail": projection.get("audit_trail", [])
    }
